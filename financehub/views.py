
import os
import tempfile
import unicodedata
import datetime  # This imports the module (use datetime.datetime, datetime.date, etc.)
import json
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.core.paginator import Paginator
from django.db.models import Q
from django.core.cache import cache
from django.db.models import Sum

import boto3
import tempfile
import os
import uuid


from .models import (
    UploadHistory,
    Lcc,
    Feedback,
    ExecutiveVisitScheduling,
    Clu,
    Freshdesk,
    DueNotice,
    Visiter,
    Dialer, # ✅ ADD THIS
)

# Forms
from .forms import FeedbackForm
from django.conf import settings
# Celery tasks
from financehub.tasks import (
    process_universal_file,
)


from io import BytesIO
import pandas as pd
import datetime
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Font

from .models import EmployeeMaster, Clu


# Decorator for session protection
from django.contrib.auth.decorators import login_required

# ---------------------------------------------------------------------
# LOGIN
# ---------------------------------------------------------------------
def fh_login(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            request.session["financehub_user"] = user.id
            return redirect("upload_loan_data")
        else:
            messages.error(request, "Invalid username or password")

    return render(request, "financehub/login.html")


# ---------------------------------------------------------------------
# LOGOUT
# ---------------------------------------------------------------------
def fh_logout(request):
    request.session.pop("financehub_user", None)
    return redirect("fh_login")


# ---------------------------------------------------------------------
# SESSION CHECK DECORATOR
# ---------------------------------------------------------------------
def financehub_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get("financehub_user"):
            return redirect("fh_login")
        return view_func(request, *args, **kwargs)
    return wrapper


MAX_UPLOAD_SIZE = 25 * 1024 * 1024  # 25 MB


# ---------------------------------------------------------------------
# FILE UPLOAD + CELERY PROCESSING
# ---------------------------------------------------------------------
# ---------------------------------------------------------------------
# UNIVERSAL FILE UPLOAD (WITH DROPDOWN)
# ---------------------------------------------------------------------
from financehub.tasks import process_universal_file
from .models import UploadHistory

FILE_TYPES = [
    ("lcc", "LCC"),
    ("collection_allocations", "Collection Allocations"),
    ("clu", "CLU"),
    ("repo", "Repo"),
    ("paid", "Paid"),
    ("closed", "Closed"),
    ("dialer", "Dialer"),
    ("duenotice", "Due Notice"),
    ("visiter", "Visiter"),
    ("employee_master", "Employee Master"),
    ("freshdesk", "Freshdesk"),
    ("esebuzz", "EseBuzz"),
    ("hero", "Hero"),
    ("kotakecs", "Kotak ECS"),
    ("smsquare", "SMSquare"),
    ("upi", "UPI"),
    ("executive_visit_scheduling", "Executive Visit Scheduling"),



]

@financehub_required
def upload_loan_data(request):

    msg = None
    error = None

    if request.method == "POST":

        file_type = request.POST.get("file_type")
        file = request.FILES.get("file")

        if not file_type:
            return render(request, "financehub/upload.html",
                          {"error": "Please select file type.", "file_types": FILE_TYPES})

        if not file:
            return render(request, "financehub/upload.html",
                          {"error": "Please choose a file.", "file_types": FILE_TYPES})

        ext = file.name.split(".")[-1].lower()
        if ext not in ("csv", "xlsx", "xls"):
            return render(request, "financehub/upload.html",
                          {"error": "Only CSV / XLS / XLSX allowed.", "file_types": FILE_TYPES})

        # Create upload record with file saved to S3 (using finance_uploads field)
        upload = UploadHistory.objects.create(
            filename=file.name,
            uploaded_by=request.user.username,
            file_type=file_type,
            finance_uploads=file,  # This saves to S3 in 'finance_uploads/' folder
            status="processing",
            total_rows=0,
            processed_rows=0
        )

        # Download the file from S3 to temp location for Celery processing

        s3 = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME,
        )

        # Create unique temp file path
        tmp_dir = tempfile.gettempdir()
        tmp_path = os.path.join(
            tmp_dir,
            f"{uuid.uuid4()}_{file.name}"
        )

        # Download from S3 to temp location
        s3.download_file(
            settings.AWS_STORAGE_BUCKET_NAME,
            upload.finance_uploads.name,
            tmp_path
        )

        # Launch celery task
        process_universal_file.delay(upload.id, tmp_path, ext, file_type)

        msg = f"Upload started! Upload ID = {upload.id}"

        return render(request, "financehub/upload.html", {
            "msg": msg,
            "file_types": FILE_TYPES,
            "upload_id": upload.id
        })

    return render(request, "financehub/upload.html", {
        "file_types": FILE_TYPES,
        "ADMIN_USER": settings.ADMIN_USER
    })

# LCC LIST WITH POWER SEARCH + PAGINATION (100 PER PAGE)
from django.db.models import Q
from django.core.paginator import Paginator


# ---------------------------------------------------------------------
# LCC LIST WITH PAGINATION + SEARCH
# ---------------------------------------------------------------------
import unicodedata
from django.db.models import Q
from django.core.paginator import Paginator

import unicodedata
from django.db.models import Q
from django.core.paginator import Paginator

def normalize_excel_text(text):
    """Clean any hidden characters & normalize Excel-pasted values."""
    if not text:
        return ""

    # Normalize Unicode
    text = unicodedata.normalize("NFKD", text)

    # Remove invisible characters including TAB (%09)
    INVISIBLE = ["\u200b", "\u200c", "\u200d", "\ufeff", "\t", "\n", "\r"]
    for ch in INVISIBLE:
        text = text.replace(ch, "")

    # Normalize various hyphens
    HYPHENS = ["‐", "-", "‒", "–", "—", "―"]
    for h in HYPHENS:
        text = text.replace(h, "-")

    # Keep only alphanumeric and hyphen
    cleaned = []
    for c in text:
        if c.isalnum() or c == "-":
            cleaned.append(c)

    return "".join(cleaned).strip()



import boto3
from django.conf import settings
from django.db.models import Q
from django.core.paginator import Paginator
from django.shortcuts import render
from django.core.cache import cache


# -----------------------------------------------------
# ✅ S3 CLIENT
# -----------------------------------------------------
s3 = boto3.client(
    "s3",
    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
    region_name=settings.AWS_S3_REGION_NAME,
)


# -----------------------------------------------------
# ✅ FETCH ALL NOTICE FILES (WITH PAGINATION + FIXES)
# -----------------------------------------------------
def get_all_notice_files():

    file_map = {}
    continuation_token = None

    try:
        while True:
            params = {
                "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
                "Prefix": "notices/"
            }

            if continuation_token:
                params["ContinuationToken"] = continuation_token

            response = s3.list_objects_v2(**params)

            for obj in response.get("Contents", []):
                key = obj["Key"]

                # filename extraction
                filename = key.split("/")[-1].lower().strip()

                # normalize filename
                filename = filename.replace(" ", "_")

                if "_" not in filename:
                    continue

                # safe parsing
                parts = filename.split("_", 1)
                loan_number = parts[0].upper()
                file_type = parts[1].replace(".pdf", "")

                # init map
                if loan_number not in file_map:
                    file_map[loan_number] = {
                        "borrower": None,
                        "guarantor": None,
                        "co_borrower": None
                    }

                # 🔐 secure URL (IMPORTANT)
                url = s3.generate_presigned_url(
                    "get_object",
                    Params={
                        "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
                        "Key": key
                    },
                    ExpiresIn=3600
                )

                # exact matching (no partial bugs)
                if file_type == "co_borrower":
                    file_map[loan_number]["co_borrower"] = url

                elif file_type == "guarantor":
                    file_map[loan_number]["guarantor"] = url

                elif file_type == "borrower":
                    file_map[loan_number]["borrower"] = url

            # pagination check
            if response.get("IsTruncated"):
                continuation_token = response.get("NextContinuationToken")
            else:
                break

    except Exception as e:
        print("S3 ERROR:", e)
        return {}

    return file_map


# -----------------------------------------------------
# ✅ CACHE WRAPPER (VERY IMPORTANT 🚀)
# -----------------------------------------------------
def get_all_notice_files_cached():
    data = cache.get("notice_files")

    if not data:
        data = get_all_notice_files()
        cache.set("notice_files", data, timeout=300)  # 5 minutes

    return data


# -----------------------------------------------------
# ✅ MAIN VIEW
# -----------------------------------------------------
@financehub_required
def lcc_list(request):

    search_raw = request.GET.get("search", "")
    search_clean = normalize_excel_text(search_raw).strip()

    base_qs = Lcc.objects.all()

    # -----------------------------------------------------
    # ✅ NO SEARCH
    # -----------------------------------------------------
    if not search_clean:
        qs = base_qs.order_by("id")

    else:
        # -----------------------------------------------------
        # ✅ EXACT MATCH
        # -----------------------------------------------------
        primary = base_qs.filter(
            Q(loan_number__iexact=search_clean) |
            Q(cust_mobile__iexact=search_clean) |
            Q(guarantor_mobile__iexact=search_clean) |
            Q(vehicle_no__iexact=search_clean)
        )

        if primary.exists():

            mobile_set = set(
                x for x in primary.values_list("cust_mobile", flat=True)
                if x not in ["", None, "0"]
            ) | set(
                x for x in primary.values_list("guarantor_mobile", flat=True)
                if x not in ["", None, "0"]
            )

            vehicle_set = set(
                x for x in primary.values_list("vehicle_no", flat=True)
                if x not in ["", None]
            )

            qs = base_qs.filter(
                Q(cust_mobile__in=mobile_set) |
                Q(guarantor_mobile__in=mobile_set) |
                Q(vehicle_no__in=vehicle_set) |
                Q(id__in=primary.values("id"))
            ).distinct().order_by("id")

        else:
            # -----------------------------------------------------
            # ✅ NAME SEARCH
            # -----------------------------------------------------
            qs = base_qs.filter(
                Q(customer_name__icontains=search_clean) |
                Q(guarantor__icontains=search_clean)
            ).order_by("id")

    # -----------------------------------------------------
    # ✅ PAGINATION
    # -----------------------------------------------------
    paginator = Paginator(qs, 50)
    page = request.GET.get("page")
    page_obj = paginator.get_page(page)

    # -----------------------------------------------------
    # ✅ LOAD S3 FILES (CACHED 🚀)
    # -----------------------------------------------------
    file_map = get_all_notice_files_cached()

    # -----------------------------------------------------
    # ✅ ATTACH PDF LINKS
    # -----------------------------------------------------
    for obj in page_obj:
        loan_number = (obj.loan_number or "").strip().upper()

        files = file_map.get(loan_number, {})

        obj.borrower_pdf = files.get("borrower")
        obj.guarantor_pdf = files.get("guarantor")
        obj.co_borrower_pdf = files.get("co_borrower")

    # -----------------------------------------------------
    # ✅ QUERY STRING
    # -----------------------------------------------------
    params = request.GET.copy()
    params.pop("page", None)
    params["search"] = search_clean

    return render(request, "financehub/lcc.html", {
        "data": page_obj,
        "search": search_clean,
        "query_string": params.urlencode(),
    })


def build_latest_payment_map(loan_numbers):
    payment_latest_map = {}
    payment_source_map = {}

    def push_payment(loan, status, date, amount, source):
        date_parsed = parse_payment_date_safe(date)
        if not loan:
            return

        # track all sources
        payment_source_map.setdefault(loan, set()).add(source)

        # date required to be latest
        if not date_parsed:
            return

        current = payment_latest_map.get(loan)
        if not current or date_parsed > current["date"]:
            payment_latest_map[loan] = {
                "status": clean_payment_value(status),
                "date": date_parsed,
                "amount": clean_payment_value(amount),
                "source": source,  # temp
            }

    # HERO
    for h in Hero.objects.filter(referencenumber__in=loan_numbers):
        push_payment(h.referencenumber, h.status, h.date, h.amount, "HERO")

    # KOTAK
    for k in KotakECS.objects.filter(loannumber__in=loan_numbers):
        push_payment(k.loannumber, k.ecsstatus, k.ecsdate, k.amount, "KOTAK")

    # ESEBUZZ
    for e in EseBuzz.objects.filter(loanno__in=loan_numbers):
        push_payment(e.loanno, e.status, e.initiateddate, e.amount, "ESEBUZZ")

    # SMSQUARE
    for s in Smsquare.objects.filter(uniqueregistrationnumber__in=loan_numbers):
        push_payment(s.uniqueregistrationnumber, s.status, s.date, s.amount, "SMSQUARE")

    # UPI
    for u in Upi.objects.filter(loannoreference__in=loan_numbers):
        push_payment(u.loannoreference, u.paymentstatus, u.paymentdatetime, u.transactionamount, "UPI")

    # merge sources
    for loan, latest in payment_latest_map.items():
        latest["source"] = " + ".join(sorted(payment_source_map.get(loan, [])))

    return payment_latest_map


# ---------------------------------------------------------------------
# CREATE FEEDBACK
# ---------------------------------------------------------------------
def normalize_mobile(num):
    if not num:
        return None
    s = str(num).strip()
    if s.startswith("91") and len(s) == 12:
        return s[2:]
    return s

@financehub_required
def feedback_create(request):

    loan_no = request.GET.get("loan", "")
    emp_id = request.user.username

    # ----------------------------
    # CASE 1: DIRECT OPEN
    # ----------------------------
    if not loan_no:

        if request.method == "POST":
            form = FeedbackForm(request.POST)
            if form.is_valid():
                form.save()
                return redirect("feedback_list")
        else:
            form = FeedbackForm(initial={"EmpID": emp_id})

        return render(request, "financehub/feedback_form.html", {
            "form": form,
            "loan_no": "",
            "combined_rows": [],
            "freshdesk_tickets": [],   # ✅ SAFE DEFAULT
        })

    # ----------------------------
    # CASE 2: OPENED FROM LCC TABLE
    # ----------------------------
    try:
        l = Lcc.objects.get(loan_number=loan_no)
    except Lcc.DoesNotExist:
        l = None

    payment = None
    if loan_no:
        payment_map = build_latest_payment_map([loan_no])
        payment = payment_map.get(loan_no)


    cust_mobile = l.cust_mobile if l else ""
    guar_mobile = l.guarantor_mobile if l else ""
    veh_no = l.vehicle_no if l else ""
    cust_name = l.customer_name if l else ""
    guar_name = l.guarantor if l else ""

    # =========================================================
    # DIALER DATA (LATEST PER DISP)
    # =========================================================


    dialer_row = {
        "Dialer_PTP": "",
        "Dialer_PTP_Date": "",
        "Dialer_PTP_Remarks": "",

        "Dialer_RTP": "",
        "Dialer_RTP_Date": "",
        "Dialer_RTP_Remarks": "",

        "Dialer_Thirdparty": "",
        "Dialer_Thirdparty_Date": "",
        "Dialer_Thirdparty_Remarks": "",

        "Dialer_other": "",
        "Dialer_other_Date": "",
        "Dialer_other_Remarks": "",
    }

    cust_mobile_norm = normalize_mobile(cust_mobile)

    if cust_mobile_norm:
        base_qs = Dialer.objects.filter(
            mobile=cust_mobile_norm
        ).exclude(ptp_date__isnull=True).exclude(ptp_date="")

        # PTP
        obj = base_qs.filter(disp__iexact="PTP").order_by("-created_at").first()
        if obj:
            dialer_row.update({
                "Dialer_PTP": obj.disp,
                "Dialer_PTP_Date": obj.ptp_date,
                "Dialer_PTP_Remarks": obj.remarks,
            })

        # RTP
        obj = base_qs.filter(disp__iexact="RTP").order_by("-created_at").first()
        if obj:
            dialer_row.update({
                "Dialer_RTP": obj.disp,
                "Dialer_RTP_Date": obj.ptp_date,
                "Dialer_RTP_Remarks": obj.remarks,
            })

        # THIRD PARTY
        obj = base_qs.filter(disp__iexact="THIRD PARTY").order_by("-created_at").first()
        if obj:
            dialer_row.update({
                "Dialer_Thirdparty": obj.disp,
                "Dialer_Thirdparty_Date": obj.ptp_date,
                "Dialer_Thirdparty_Remarks": obj.remarks,
            })

        # OTHER
        obj = base_qs.exclude(
            disp__in=["PTP", "RTP", "THIRD PARTY"]
        ).order_by("-created_at").first()
        if obj:
            dialer_row.update({
                "Dialer_other": obj.disp,
                "Dialer_other_Date": obj.ptp_date,
                "Dialer_other_Remarks": obj.remarks,
            })


    # =========================================================
    # ✅ NEW: FRESHDESK MATCH (LOAN NUMBER INSIDE SUBJECT)
    # =========================================================
    freshdesk_tickets = []

    if loan_no:
        freshdesk_tickets = Freshdesk.objects.filter(
            subject__icontains=loan_no
        ).order_by("-created_at")
    # =========================================================

    duenotices = []
    if loan_no:
        duenotices = DueNotice.objects.filter(
            loan_number=loan_no
        ).order_by("-id")

    # =========================================================
# ✅ CLU VISITS (MATCHED BY LOAN NUMBER)
# =========================================================
    clu_visits = []

    if loan_no:
        clu_visits = Clu.objects.filter(
            loan_number=loan_no
        ).order_by("-created_at")

    # VISITERS (MATCH BY LOAN NUMBER)
    visitors = []
    if loan_no:
        visitors = Visiter.objects.filter(
            loan_number=loan_no
        ).order_by("-created_at")



    # ----------------------------
    # EXACT LOAN FEEDBACK
    # ----------------------------
    exact_fb = list(
        Feedback.objects.filter(LoanNO=loan_no).order_by("-id")
    )

    # ----------------------------
    # RELATED LOANS (SAFE LOGIC)
    # ----------------------------
    filters = Q()

    if cust_mobile not in ["", None, "0"]:
        filters |= Q(cust_mobile=cust_mobile)

    if guar_mobile not in ["", None, "0"]:
        filters |= Q(guarantor_mobile=guar_mobile)

    if veh_no not in ["", None]:
        filters |= Q(vehicle_no=veh_no)

    if filters:
        related_qs = Lcc.objects.filter(filters).exclude(loan_number=loan_no)
    else:
        related_qs = Lcc.objects.none()

    related_loan_numbers = [x.loan_number for x in related_qs]

    # ----------------------------
    # RELATED FEEDBACK
    # ----------------------------
    related_fb_all = Feedback.objects.filter(
        LoanNO__in=related_loan_numbers
    ).order_by("-id")

    related_map = {}
    for fb in related_fb_all:
        related_map.setdefault(fb.LoanNO, []).append(fb)

    # ----------------------------
    # BUILD COMBINED HISTORY TABLE
    # ----------------------------
    combined_rows = []

    for fb in exact_fb:
        combined_rows.append({
            "loan": fb.LoanNO,
            "vehicle": veh_no,
            "cust_mobile": cust_mobile,
            "guar_mobile": guar_mobile,
            "cust_name": cust_name,
            "guar_name": guar_name,
            "feedback": fb,
        })

    for rl in related_qs:
        fblist = related_map.get(rl.loan_number, [])
        if fblist:
            for fb in fblist:
                combined_rows.append({
                    "loan": rl.loan_number,
                    "vehicle": rl.vehicle_no,
                    "cust_mobile": rl.cust_mobile,
                    "guar_mobile": rl.guarantor_mobile,
                    "cust_name": rl.customer_name,
                    "guar_name": rl.guarantor,
                    "feedback": fb,
                })
        else:
            combined_rows.append({
                "loan": rl.loan_number,
                "vehicle": rl.vehicle_no,
                "cust_mobile": rl.cust_mobile,
                "guar_mobile": rl.guarantor_mobile,
                "cust_name": rl.customer_name,
                "guar_name": rl.guarantor,
                "feedback": None,
            })

    # ----------------------------
    # FORM
    # ----------------------------
    initial = {
        "EmpID": emp_id,
        "LoanNO": loan_no,
        "customer_name": cust_name,
        "vehicle_no": veh_no,
    }

    if request.method == "POST":
        form = FeedbackForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("feedback_list")
    else:
        form = FeedbackForm(initial=initial)

    return render(request, "financehub/feedback_form.html", {
        "form": form,
        "loan_no": loan_no,
        "combined_rows": combined_rows,
        "freshdesk_tickets": freshdesk_tickets,
        "duenotices": duenotices,
        "clu_visits": clu_visits,  # ✅ NEW
        "visitors": visitors,  # ✅ NEW
        "payment": payment,
        "dialer_row": dialer_row,

    })


from django.shortcuts import render
from django.db.models import Q
import datetime
from .models import Feedback


# ---------------------------------------------------------------------
# FEEDBACK LIST - POWER SEARCH + PAGINATION (100/page)
# ---------------------------------------------------------------------

from django.core.paginator import Paginator
from django.db.models import Q
import datetime

@financehub_required
def feedback_list(request):

    def clean(v):
        return v.strip() if v and v != "None" else None

    emp       = clean(request.GET.get("emp"))
    date_str  = clean(request.GET.get("date"))
    ftype     = clean(request.GET.get("ftype"))
    ctype     = clean(request.GET.get("ctype"))
    visiting  = clean(request.GET.get("visiting"))
    executive = clean(request.GET.get("executive"))
    ptp       = clean(request.GET.get("ptp"))
    vdate     = clean(request.GET.get("vdate"))
    search    = clean(request.GET.get("search"))

    qs = Feedback.objects.all().order_by("-id")

    # Power search
    if search:
        qs = qs.filter(
            Q(LoanNO__icontains=search) |
            Q(customer_name__icontains=search) |
            Q(vehicle_no__icontains=search) |
            Q(Remarks__icontains=search) |
            Q(EmpID__icontains=search)
        )

    # Filters
    if emp:
        qs = qs.filter(EmpID__iexact=emp)

    if date_str:
        try:
            qs = qs.filter(Date=datetime.datetime.strptime(date_str, "%Y-%m-%d"))
        except:
            pass

    if ftype:
        qs = qs.filter(Dropdown=ftype)

    if ctype:
        qs = qs.filter(feedback_dropdwon=ctype)

    if visiting:
        qs = qs.filter(visiting_required=(visiting == "yes"))

    if executive:
        qs = qs.filter(executive_id=executive)

    if ptp:
        try:
            qs = qs.filter(PTPDate=datetime.datetime.strptime(ptp, "%Y-%m-%d"))
        except:
            pass

    if vdate:
        try:
            qs = qs.filter(visit_date=datetime.datetime.strptime(vdate, "%Y-%m-%d"))
        except:
            pass

    # Pagination
    paginator = Paginator(qs, 1000)
    page_obj = paginator.get_page(request.GET.get("page"))

    params = request.GET.copy()
    params.pop("page", None)

    filters = {
        "emp": emp or "",
        "date": date_str or "",
        "ftype": ftype or "",
        "ctype": ctype or "",
        "visiting": visiting or "",
        "executive": executive or "",
        "ptp": ptp or "",
        "vdate": vdate or "",
        "search": search or "",
    }

    return render(request, "financehub/feedback_list.html", {
        "data": page_obj,
        "filters": filters,
        "query_string": params.urlencode(),
        "FTYPES": Feedback.DROPDOWN_CHOICES,
        "CTYPES": Feedback.FEEDBACK_CHOICES,
    })





# ---------------------------------------------------------------
# AJAX PROGRESS CHECK
# ---------------------------------------------------------------
# financehub/views.py (top imports)
from django.http import JsonResponse
# ... other imports remain

# upload_progress (already exists), ensure it uses JsonResponse
def upload_progress(request, upload_id):
    try:
        u = UploadHistory.objects.get(id=upload_id)
        return JsonResponse({
            "status": u.status,
            "processed": u.processed_rows,
            "total": u.total_rows,
            "percent": u.progress_percentage(),
            "error": u.error_message or "",
        })
    except UploadHistory.DoesNotExist:
        return JsonResponse({"error": "Invalid Upload ID"}, status=404)






from openpyxl import Workbook
from django.http import HttpResponse


def export_lcc_excel(final_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Status Report"

    # ======================
    # HEADER ROW
    # ======================
    headers = [
        "Loan Number",
        "Customer Name",
        "Company",
        "Branch",
        "Division",
        "EMI",
        "Loan Status",
        "Received Date",
        "CM",
        "TL",
        "Executive",
        "Visit Date",
        "Visit Status",
    ]
    ws.append(headers)

    # ======================
    # DATA ROWS
    # ======================
    for r in final_data:
        ws.append([
            r["loan_number"],
            r["customer_name"],
            r["company"],
            r["branch"],
            r.get("division", ""),
            r["emi_due_2"],
            r["loan_status"],
            r["received_date"],
            r["cm"],
            r["tl"],
            r["exec"],
            r["visit_date"],
            r["visit_status"],
        ])

    # ======================
    # RESPONSE
    # ======================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="loan_status_report.xlsx"'
    )

    wb.save(response)
    return response




from django.http import HttpResponse
from reportlab.lib.pagesizes import legal, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from datetime import date

# ======================================================
# STYLES
# ======================================================
styles = getSampleStyleSheet()

wrap_style = ParagraphStyle(
    "wrap",
    parent=styles["Normal"],
    fontSize=6.5,
    leading=7.5,
    wordWrap="CJK",
)

# ======================================================
# HELPERS
# ======================================================
def clean_colon_zero(value):
    if not value:
        return ""
    value = str(value).strip()
    if value in ("00:0", "0:00"):
        return ""
    return value


def wrap_cell(value, max_chars=50):
    if not value:
        return Paragraph("", wrap_style)

    text = str(value)
    lines = [text[i:i + max_chars] for i in range(0, len(text), max_chars)]
    return Paragraph("<br/>".join(lines), wrap_style)


# ======================================================
# 🔥 FINAL SORT KEY (EMI → DATE)
# ======================================================
def lcc_final_sort_key(row):
    """
    FINAL SORT ORDER:

    1️⃣ EMI DUE (PRIMARY)
       - SEZ first
       - Numeric EMI (higher → lower)
       - Zero / blank last

    2️⃣ INSTALLMENT DATE (SECONDARY)
       - Oldest date first
    """

    # -----------------------
    # INSTALLMENT DATE
    # -----------------------
    inst_date = row.get("installment_date")
    if not inst_date:
        inst_date = date.max  # empty dates last

    # -----------------------
    # EMI PROCESSING
    # -----------------------
    emi_raw = str(row.get("emi_due_2", "")).strip().lower()

    # SEZ → TOP PRIORITY
    if emi_raw == "sez":
        emi_group = 0
        emi_value = float("inf")

    # NUMERIC EMI
    else:
        try:
            emi_value = float(emi_raw)
            if emi_value > 0:
                emi_group = 1
            else:
                emi_group = 2  # zero EMI
        except Exception:
            emi_group = 2
            emi_value = 0

    # -----------------------
    # SORT TUPLE
    # -----------------------
    return (
        emi_group,      # SEZ → numeric → zero
        -emi_value,     # BIG EMI FIRST
        inst_date       # OLDEST DATE FIRST
    )


# ======================================================
# PDF EXPORT
# ======================================================
def export_lcc_pdf(final_data):

    # ✅ APPLY FINAL SORT
    final_data = sorted(final_data, key=lcc_final_sort_key)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="lcc_legal_report.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(legal),
        leftMargin=6,
        rightMargin=6,
        topMargin=8,
        bottomMargin=8,
    )

    table_data = [[
        "S/No","Company","Branch","Loan No","Vehicle No","Loan Date",
        "Customer Name","Mobile","Guarantor","G. Mobile",
        "Class","Inst Date","Month TBC","Total",
        "LPC","VAS","EMI","EMI 2","Run EMI",
        "Paid","Bal","Inst",
        "Last Rcvd","Seize","Address","Executive"
    ]]

    for i, r in enumerate(final_data, 1):
        table_data.append([
            i,
            wrap_cell(r.get("company"), 20),
            wrap_cell(r.get("branch"), 20),
            wrap_cell(r.get("loan_number"), 20),
            wrap_cell(r.get("vehicle_no"), 20),
            wrap_cell(r.get("loan_date"), 15),
            wrap_cell(r.get("customer_name"), 25),
            wrap_cell(r.get("cust_mobile"), 15),
            wrap_cell("" if str(r.get("guarantor")) == "0" else r.get("guarantor"), 25),
            wrap_cell("" if str(r.get("guarantor_mobile")) == "0" else r.get("guarantor_mobile"), 15),
            wrap_cell(r.get("vehicle_class"), 15),
            wrap_cell(r.get("installment_date"), 15),
            wrap_cell(clean_colon_zero(r.get("month_tbc")), 10),
            wrap_cell(r.get("total_dues"), 10),
            wrap_cell(r.get("lpc_dues"), 10),
            wrap_cell(r.get("vas_hl"), 10),
            wrap_cell(r.get("emi_due"), 10),
            wrap_cell(
                "" if str(r.get("emi_due_2")).strip().lower() == "0"
                else str(r.get("emi_due_2")).upper(),
                10
            ),
            wrap_cell(r.get("running_emi"), 10),
            wrap_cell(r.get("paid_inst"), 10),
            wrap_cell(r.get("balance_inst"), 10),
            wrap_cell(r.get("inst"), 8),
            wrap_cell(r.get("last_rcvd_date"), 15),
            wrap_cell(r.get("seize_date"), 15),
            wrap_cell(r.get("customer_address"), 50),
            wrap_cell(r.get("collection_executive"), 20),
        ])

    col_widths = [
        18, 38, 40, 48, 48, 40,
        58, 48, 58, 48,
        33, 40, 28, 38,
        28, 28, 28, 28, 28,
        18, 18, 18,
        40, 38,
        100, 60
    ]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("BACKGROUND", (0,0), (-1,0), colors.darkblue),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 6),
        ("LEFTPADDING", (0,0), (-1,-1), 3),
        ("RIGHTPADDING", (0,0), (-1,-1), 3),
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))

    doc.build([table])
    return response





import datetime
from collections import Counter
from django.shortcuts import render
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.db.models import Q


from .models import (
    Lcc,
    ExecutiveVisitScheduling,
    Clu,
    CollectionAllocations,
    Repo,
    Paid,
    Closed,
    Hero,
    KotakECS,
    EseBuzz,
    Smsquare,
    Upi,
    Freshdesk,
    Dialer,
    DueNotice,
)

def parse_emi_bucket(value):
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


EMI_BUCKETS = {
    "0_0": (0, 0),
    "1_5": (1, 5),
    "6_10": (6, 10),
    "11_20": (11, 20),
    "21_50": (21, 50),
    "50_plus": (51, None),
}

# ------------------------------------------------------------------
# CLU VISIT DATE PARSER
# ------------------------------------------------------------------
def parse_visited_on(value):
    if not value:
        return None

    value = value.strip()
    formats = [
        "%b %d,%Y, %I:%M:%S %p",
        "%d-%b-%Y %I:%M %p",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y",
    ]

    for fmt in formats:
        try:
            return datetime.datetime.strptime(value, fmt)
        except Exception:
            continue

    return None


# ------------------------------------------------------------------
# EMI NORMALIZER
# ------------------------------------------------------------------
def normalize_emi(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def parse_payment_date_safe(val):
    if not val:
        return None

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%b %d,%Y, %I:%M:%S %p",
    ):
        try:
            return datetime.datetime.strptime(str(val).strip(), fmt)
        except Exception:
            continue
    return None

def clean_payment_value(val):
    if not val:
        return None
    val = str(val).strip()
    return None if val.lower() == "nan" else val
# ------------------------------------------------------------------
# PAYMENT STATUS LOGIC (MINIMAL SAFE FIX)
# ------------------------------------------------------------------





# ------------------------------------------------------------------
# EXECUTIVE VISIT SCHEDULE LIST
# ------------------------------------------------------------------
import hashlib



# ------------------------------------------------------------------
# EXECUTIVE VISIT SCHEDULE LIST
# ------------------------------------------------------------------
# views.py
from django.db import connection
from django.core.paginator import Paginator
from django.core.cache import cache
from collections import Counter, defaultdict
from django.shortcuts import render
from django.db.models import Q
from .models import LoanStatusCache  # Add this import
from django.db.models import Q, OuterRef, Subquery,Count

from django.core.paginator import Paginator
from django.core.cache import cache
from collections import Counter, defaultdict
from django.shortcuts import render
from django.db.models import Q
from .models import Lcc, Paid, Repo, Closed, LoanStatusCache, ExecutiveVisitScheduling, CollectionAllocations, DueNotice, Visiter, Freshdesk, Dialer  # Added Freshdesk and Dialer
from django.db import connection
from django.core.paginator import Paginator
from django.core.cache import cache
from django.shortcuts import render
from django.db.models import Q, OuterRef, Subquery, Value, CharField, Case, When, Exists, Count
from django.db.models.functions import Coalesce
from django.db import models
from django.db import connection
from django.core.paginator import Paginator
from django.core.cache import cache
from django.shortcuts import render
from django.db.models import Q, Count, OuterRef, Exists


from django.core.paginator import Paginator
from django.core.cache import cache
from django.shortcuts import render
from django.db.models import Q, Count, OuterRef, Exists
from .models import Lcc, LoanStatusCache, ExecutiveVisitScheduling, CollectionAllocations, DueNotice, Visiter, Paid, Freshdesk, Dialer



TOLERANCE = 500

def get_loan_status(lcc, paid_amount, repo_set, closed_set):
    loan_no = lcc.loan_number

    # 1️⃣ CLOSED
    if loan_no in closed_set:
        return "CLOSED"

    # 2️⃣ REPO
    if loan_no in repo_set:
        return "REPO"

    # 3️⃣ EMI = 0 / SEZ / NONE
    if str(lcc.emi_due_2).strip().lower() in ("0", "", "none", "sez"):
        return "PAID" if paid_amount and paid_amount > 0 else "NOT_PAID"

    # 4️⃣ Parse values
    try:
        received = float(paid_amount or 0)
        month_tbc = float(lcc.month_tbc or 0)
        total_dues = float(lcc.total_dues or 0)
        emi_due_2 = float(lcc.emi_due_2)
    except Exception:
        return "NOT_PAID"

    # 5️⃣ Calculate revised month TBC
    if month_tbc == 0:
        revised_month_tbc = total_dues / emi_due_2 if emi_due_2 else 0
    else:
        revised_month_tbc = month_tbc

    # 6️⃣ FINAL DECISION
    if received > 0 and (
        received >= revised_month_tbc or
        (revised_month_tbc - received) <= TOLERANCE
    ):
        return "PAID"
    elif received > 0:
        return "PARTLY_PAID"
    else:
        return "NOT_PAID"


@financehub_required
def executive_visit_schedule_list(request):
    from django.core.paginator import Paginator
    from django.core.cache import cache
    from django.db.models import Q, Sum
    from collections import defaultdict
    import hashlib
    import json
    import time

    start_time = time.time()

    # ==========================================================
    # GET PARAMS
    # ==========================================================
    division = request.GET.get("division", "").strip()
    blc_case = request.GET.get("blc_case", "").strip()
    loanno = request.GET.get("loanno", "").strip()
    empid = request.GET.get("empid", "").strip()
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()
    role = request.GET.get("role", "").strip()
    search_empid = request.GET.get("search_empid", "").strip()
    loan_status_filter = request.GET.get("loan_status", "").strip()
    emi_bucket = request.GET.get("emi_bucket", "").strip()
    remove_zeros = request.GET.get("remove_zeros") == "1"
    remove_sez = request.GET.get("remove_sez") == "1"
    branch = request.GET.get("branch", "").strip()
    centre_name = request.GET.get("centre_name", "").strip()
    visit_filter = request.GET.get("visit_filter", "").strip()
    company = request.GET.get("company", "").strip()
    type_of_notice = request.GET.get("type_of_notice", "").strip()
    sort_field = request.GET.get("sort_field", "loan_number")
    sort_order = request.GET.get("sort_order", "asc")
    page = int(request.GET.get("page", 1))
    login_empid = request.user.username.strip()

    def valid_filter(v):
        return v and v not in ["undefined", "null", "None", ""]

    # ==========================================================
    # BUILD QUERYSET
    # ==========================================================
    qs = Lcc.objects.all().only(
        'loan_number', 'customer_name', 'vehicle_no', 'cust_mobile',
        'company', 'division', 'branch', 'centre_name', 'blc_cases',
        'emi_due_2', 'emi_due', 'month_tbc', 'total_dues', 'id'
    )

    # Apply all filters (your existing filters)
    if valid_filter(division):
        qs = qs.filter(division=division)
    if valid_filter(branch):
        qs = qs.filter(branch=branch)
    if valid_filter(centre_name):
        qs = qs.filter(centre_name=centre_name)
    if valid_filter(loanno):
        qs = qs.filter(loan_number__icontains=loanno)
    if valid_filter(company):
        qs = qs.filter(company=company)
    if valid_filter(blc_case):
        qs = qs.filter(blc_cases=blc_case)

    # ✅ FIX 1: Apply NOTICE FILTER EARLY (before status calculation)
    if valid_filter(type_of_notice):
        notice_loans = DueNotice.objects.filter(
            type_of_notice__iexact=type_of_notice
        ).values_list('loan_number', flat=True).distinct()
        qs = qs.filter(loan_number__in=notice_loans)

    # EMI Bucket
    if valid_filter(emi_bucket):
        from django.db.models import FloatField
        from django.db.models.functions import Cast

        qs = qs.annotate(emi_float=Cast('emi_due_2', FloatField()))

        if emi_bucket == "0_0":
            qs = qs.filter(emi_float=0)
        elif emi_bucket == "1_5":
            qs = qs.filter(emi_float__gte=1, emi_float__lte=5)
        elif emi_bucket == "6_10":
            qs = qs.filter(emi_float__gte=6, emi_float__lte=10)
        elif emi_bucket == "11_20":
            qs = qs.filter(emi_float__gte=11, emi_float__lte=20)
        elif emi_bucket == "21_50":
            qs = qs.filter(emi_float__gte=21, emi_float__lte=50)
        elif emi_bucket == "50_plus":
            qs = qs.filter(emi_float__gt=50)

    if remove_zeros:
        qs = qs.exclude(emi_due_2__in=["0", "0.0"])
    if remove_sez:
        qs = qs.exclude(emi_due_2__iexact="sez")

    # Role filter
    if role == "CM":
        cm_loans = CollectionAllocations.objects.filter(
            manager_employee_id__iexact=login_empid
        ).values_list('loan_number', flat=True)
        qs = qs.filter(loan_number__in=cm_loans)
        if valid_filter(search_empid):
            search_loans = CollectionAllocations.objects.filter(
                Q(tl_employee_id__iexact=search_empid) | Q(employee_id__iexact=search_empid)
            ).values_list('loan_number', flat=True)
            qs = qs.filter(loan_number__in=search_loans)
    elif role == "TL":
        tl_loans = CollectionAllocations.objects.filter(
            tl_employee_id__iexact=login_empid
        ).values_list('loan_number', flat=True)
        qs = qs.filter(loan_number__in=tl_loans)
        if valid_filter(search_empid):
            search_loans = CollectionAllocations.objects.filter(
                tl_employee_id__iexact=search_empid
            ).values_list('loan_number', flat=True)
            qs = qs.filter(loan_number__in=search_loans)
    elif role == "EXEC":
        exec_loans = CollectionAllocations.objects.filter(
            employee_id__iexact=login_empid
        ).values_list('loan_number', flat=True)
        qs = qs.filter(loan_number__in=exec_loans)
        if valid_filter(search_empid):
            qs = qs.filter(loan_number__in=CollectionAllocations.objects.filter(
                employee_id__iexact=search_empid
            ).values_list('loan_number', flat=True))

    # Visit filters
    if valid_filter(empid):
        visit_loans = ExecutiveVisitScheduling.objects.filter(
            empid__icontains=empid
        ).values_list('loanno', flat=True)
        qs = qs.filter(loan_number__in=visit_loans)

    if from_date and to_date:
        visit_loans = ExecutiveVisitScheduling.objects.filter(
            visit_schedule_date__range=[from_date, to_date]
        ).values_list('loanno', flat=True)
        qs = qs.filter(loan_number__in=visit_loans)
    elif from_date:
        visit_loans = ExecutiveVisitScheduling.objects.filter(
            visit_schedule_date__gte=from_date
        ).values_list('loanno', flat=True)
        qs = qs.filter(loan_number__in=visit_loans)
    elif to_date:
        visit_loans = ExecutiveVisitScheduling.objects.filter(
            visit_schedule_date__lte=to_date
        ).values_list('loanno', flat=True)
        qs = qs.filter(loan_number__in=visit_loans)

    if visit_filter == "visited":
        visit_loans = ExecutiveVisitScheduling.objects.filter(
            visit_status__iexact="visited"
        ).values_list('loanno', flat=True)
        qs = qs.filter(loan_number__in=visit_loans)
    elif visit_filter == "not_visited":
        visit_loans = ExecutiveVisitScheduling.objects.filter(
            visit_status__iexact="not_visited"
        ).values_list('loanno', flat=True)
        qs = qs.filter(loan_number__in=visit_loans)
    elif visit_filter == "scheduled":
        visit_loans = ExecutiveVisitScheduling.objects.all().values_list('loanno', flat=True)
        qs = qs.filter(loan_number__in=visit_loans)
    elif visit_filter == "not_scheduled":
        visit_loans = ExecutiveVisitScheduling.objects.all().values_list('loanno', flat=True)
        qs = qs.exclude(loan_number__in=visit_loans)

    # ==========================================================
    # GET TOTAL COUNT (BEFORE LOAN STATUS FILTER)
    # ==========================================================
    total_before_status = qs.count()

    if total_before_status == 0:
        dropdowns = cache.get('dropdowns_final_v4')
        if not dropdowns:
            dropdowns = {
                'branches': [], 'centres': [], 'notice_types': [], 'blc_cases': [], 'companies': []
            }
        context = {
            "data": [], "page_obj": None, "total_count": 0,
            "status_counts": {'CLOSED':0, 'REPO':0, 'PAID':0, 'PARTLY_PAID':0, 'NOT_PAID':0},
            **dropdowns
        }
        return render(request, "financehub/executive_visit_schedule_list.html", context)

    # ==========================================================
    # ✅ FIX 2: GET ALL LOAN NUMBERS FOR STATUS CALCULATION
    # ==========================================================
    all_filtered_loans = list(qs.values_list('loan_number', flat=True))

    # Get closed and repo in bulk
    closed_dict = set(Closed.objects.filter(loan_number__in=all_filtered_loans).values_list('loan_number', flat=True))
    repo_dict = set(Repo.objects.filter(agreement_number__in=all_filtered_loans).values_list('agreement_number', flat=True))

    # Get paid amounts in bulk
    paid_dict = dict(
        Paid.objects.filter(loan_number__in=all_filtered_loans)
        .values('loan_number')
        .annotate(total=Sum('received_amount'))
        .values_list('loan_number', 'total')
    )

    # Get LCC data in bulk
    lcc_dict = {
        l.loan_number: l for l in Lcc.objects.filter(loan_number__in=all_filtered_loans)
        .only('loan_number', 'emi_due_2', 'month_tbc', 'total_dues')
    }

    # Calculate status for each loan
    loan_status_map = {}
    status_counts = {'CLOSED': 0, 'REPO': 0, 'PAID': 0, 'PARTLY_PAID': 0, 'NOT_PAID': 0}

    for loan_num in all_filtered_loans:
        lcc_obj = lcc_dict.get(loan_num)
        if not lcc_obj:
            continue

        paid_amount = float(paid_dict.get(loan_num, 0) or 0)
        status = get_loan_status(lcc_obj, paid_amount, repo_dict, closed_dict)
        loan_status_map[loan_num] = status
        status_counts[status] += 1

    # ✅ FIX 3: APPLY LOAN STATUS FILTER (BEFORE PAGINATION)
    if valid_filter(loan_status_filter):
        filtered_loan_numbers = [loan for loan, status in loan_status_map.items() if status == loan_status_filter]
        qs = qs.filter(loan_number__in=filtered_loan_numbers)

        # Recalculate total after status filter
        total_count = len(filtered_loan_numbers)
    else:
        total_count = total_before_status

    # ==========================================================
    # PAGINATION - Get current page data
    # ==========================================================
    # Get all loan numbers for pagination
    all_loan_numbers = list(qs.values_list('loan_number', flat=True))

    # Create paginator
    paginator = Paginator(all_loan_numbers, 200)
    page_obj = paginator.get_page(page)
    current_page_loans = page_obj.object_list

    # ==========================================================
    # FETCH DATA FOR CURRENT PAGE ONLY
    # ==========================================================
    page_lcc = {l.loan_number: l for l in Lcc.objects.filter(loan_number__in=current_page_loans)}
    closed_set = set(Closed.objects.filter(loan_number__in=current_page_loans).values_list('loan_number', flat=True))
    repo_set = set(Repo.objects.filter(agreement_number__in=current_page_loans).values_list('agreement_number', flat=True))

    paid_aggregate = dict(
        Paid.objects.filter(loan_number__in=current_page_loans)
        .values('loan_number')
        .annotate(total=Sum('received_amount'))
        .values_list('loan_number', 'total')
    )

    all_visits = {}
    for v in ExecutiveVisitScheduling.objects.filter(loanno__in=current_page_loans).order_by('loanno', '-visit_schedule_date'):
        if v.loanno not in all_visits:
            all_visits[v.loanno] = v

    all_allocs = {a.loan_number: a for a in CollectionAllocations.objects.filter(loan_number__in=current_page_loans)}

    all_notices = {}
    for n in DueNotice.objects.filter(loan_number__in=current_page_loans).order_by('loan_number', '-notice_date', '-id'):
        if n.loan_number not in all_notices:
            all_notices[n.loan_number] = n

    all_received = {}
    for p in Paid.objects.filter(loan_number__in=current_page_loans, received_date__isnull=False).order_by('loan_number', '-received_date'):
        if p.loan_number not in all_received:
            all_received[p.loan_number] = p.received_date

    all_visitors = {}
    for v in Visiter.objects.filter(loan_number__in=current_page_loans).order_by('loan_number', '-created_at'):
        if v.loan_number not in all_visitors:
            all_visitors[v.loan_number] = v

    all_freshdesk = {}
    for loan_num in current_page_loans[:100]:
        fd_list = Freshdesk.objects.filter(subject__icontains=loan_num).order_by('-created_time')[:1]
        if fd_list:
            all_freshdesk[loan_num] = fd_list[0]

    all_dialer = {}
    all_mobiles = {}
    for lcc_obj in page_lcc.values():
        if lcc_obj.cust_mobile:
            all_mobiles[lcc_obj.cust_mobile] = lcc_obj.loan_number
    if all_mobiles:
        dialer_list = Dialer.objects.filter(mobile__in=list(all_mobiles.keys())).order_by('mobile', '-created_at')
        seen_mobiles = set()
        for dialer in dialer_list:
            if dialer.mobile not in seen_mobiles:
                seen_mobiles.add(dialer.mobile)
                loan_num = all_mobiles.get(dialer.mobile)
                if loan_num:
                    all_dialer[loan_num] = dialer

    # ==========================================================
    # BUILD DATA FOR CURRENT PAGE
    # ==========================================================
    notice_status_map = dict(DueNotice.NoticeStatus.choices)
    all_rows = []

    for loan_num in current_page_loans:
        lcc_obj = page_lcc.get(loan_num)
        if not lcc_obj:
            continue

        # Get status from pre-calculated map
        status = loan_status_map.get(loan_num, "NOT_PAID")

        visit = all_visits.get(loan_num)
        alloc = all_allocs.get(loan_num)
        notice = all_notices.get(loan_num)
        visitor = all_visitors.get(loan_num)
        freshdesk = all_freshdesk.get(loan_num)
        dialer = all_dialer.get(loan_num)

        all_rows.append({
            "obj": visit,
            "id": visit.id if visit else 0,
            "loan_number": lcc_obj.loan_number,
            "customer_name": lcc_obj.customer_name or "",
            "vehicle_no": lcc_obj.vehicle_no or "",
            "cust_mobile": lcc_obj.cust_mobile or "",
            "company": lcc_obj.company or "",
            "division": lcc_obj.division or "",
            "branch": lcc_obj.branch or "",
            "blc_case": lcc_obj.blc_cases or "",
            "bucket_position": lcc_obj.emi_due_2 or "",
            "loan_status": status,
            "has_schedule": bool(visit),
            "visit_date": visit.visit_schedule_date if visit else "",
            "visit_status": visit.visit_status if visit else "",
            "not_visited_reason": visit.not_visited_reason if visit else "",
            "cm": alloc.cm if alloc else "",
            "tl": alloc.tl if alloc else "",
            "exec": alloc.executive_name if alloc else "",
            "empid": visit.empid if visit else "",
            "latest_visited_on": visit.visit_schedule_date if visit else "",
            "received_date": all_received.get(loan_num, ""),
            "notice_send_to": notice.send_to if notice else "",
            "notice_bar_number": notice.bar_number if notice else "",
            "notice_date": notice.notice_date if notice else "",
            "notice_type": notice.type_of_notice if notice else "",
            "notice_status": notice.notice_status if notice else "",
            "notice_status_label": notice_status_map.get(notice.notice_status, "") if notice else "",
            "notice_delivery_date": notice.delivery_date if notice else "",
            "notice_return_date": notice.return_date if notice else "",
            "visitor_purpose": visitor.purpose if visitor else "",
            "visitor_remark": visitor.remarks if visitor else "",
            "Freshdesk_Description": freshdesk.description if freshdesk else "",
            "Freshdesk_Status": freshdesk.status if freshdesk else "",
            "Freshdesk_Group": freshdesk.group if freshdesk else "",
            "Freshdesk_Createdtime": freshdesk.created_time if freshdesk else "",
            "Dialer_PTP": dialer.ptp_date if dialer else "",
            "Dialer_PTP_Date": dialer.ptp_date if dialer else "",
            "Dialer_PTP_Remarks": dialer.remarks if dialer else "",
            "Dialer_RTP": dialer.disp if dialer else "",
            "Dialer_RTP_Date": dialer.last_received_date if dialer else "",
            "Dialer_RTP_Remarks": dialer.remarks if dialer else "",
            "Dialer_Thirdparty": dialer.executive if dialer else "",
            "Dialer_Thirdparty_Date": str(dialer.created_at)[:10] if dialer and dialer.created_at else "",
            "Dialer_Thirdparty_Remarks": dialer.customer_address if dialer else "",
            "Dialer_other": dialer.service_name if dialer else "",
            "Dialer_other_Date": dialer.call_start_time if dialer else "",
            "Dialer_other_Remarks": dialer.call_end_time if dialer else "",
            "payment_source": dialer.disp if dialer else "",
            "payment_status": "PTP" if dialer and dialer.ptp_date else "",
            "payment_date": dialer.ptp_date if dialer else "",
            "payment_amount": dialer.total_dues if dialer else "",
        })

    # Apply sorting
    reverse_sort = sort_order == "desc"

    def get_sort_key(field_name):
        def sort_key(row):
            value = row.get(field_name, "")
            if value is None or value == "":
                return ("",)
            if field_name == "id":
                return (value if value else 0,)
            elif field_name == "loan_status":
                status_order = {"PAID": 1, "PARTLY_PAID": 2, "PARTLY PAID": 2, "NOT_PAID": 3, "REPO": 4, "CLOSED": 5}
                return (status_order.get(str(value).upper(), 99),)
            elif field_name == "emi_due_2":
                try:
                    return (float(str(value).replace("SEZ", "999999").replace("sez", "999999")),)
                except:
                    return (999999,)
            else:
                return (str(value).lower(),)
        return sort_key

    all_rows.sort(key=get_sort_key(sort_field), reverse=reverse_sort)
    page_obj.object_list = all_rows

    # ==========================================================
    # GET DROPDOWNS
    # ==========================================================
    dropdowns = cache.get('dropdowns_final_v4')
    if not dropdowns:
        dropdowns = {
            'branches': list(Lcc.objects.filter(branch__isnull=False).exclude(branch='').values_list('branch', flat=True).distinct().order_by('branch')[:200]),
            'centres': list(Lcc.objects.filter(centre_name__isnull=False).exclude(centre_name='').values_list('centre_name', flat=True).distinct().order_by('centre_name')[:200]),
            'notice_types': list(DueNotice.objects.filter(type_of_notice__isnull=False).exclude(type_of_notice='').values_list('type_of_notice', flat=True).distinct().order_by('type_of_notice')[:100]),
            'blc_cases': list(Lcc.objects.filter(blc_cases__isnull=False).exclude(blc_cases='').values_list('blc_cases', flat=True).distinct().order_by('blc_cases')[:100]),
            'companies': list(Lcc.objects.filter(company__isnull=False).exclude(company='').values_list('company', flat=True).distinct().order_by('company')[:200]),
        }
        cache.set('dropdowns_final_v4', dropdowns, 3600)

    elapsed_time = time.time() - start_time
    print(f"✅ Page {page} of {paginator.num_pages} | Total: {total_count} records | Status: CLOSED={status_counts['CLOSED']}, REPO={status_counts['REPO']}, PAID={status_counts['PAID']}, PARTLY={status_counts['PARTLY_PAID']}, NOT_PAID={status_counts['NOT_PAID']} | Loaded in {elapsed_time:.2f}s")

    # ==========================================================
    # RENDER RESPONSE
    # ==========================================================
    context = {
        "data": all_rows,
        "page_obj": page_obj,
        "total_count": total_count,
        "status_counts": status_counts,
        **dropdowns,
        "division": division, "loanno": loanno, "empid": empid,
        "from_date": from_date, "to_date": to_date, "role": role,
        "search_empid": search_empid, "loan_status": loan_status_filter,
        "remove_zeros": remove_zeros, "remove_sez": remove_sez,
        "branch": branch, "centre_name": centre_name, "emi_bucket": emi_bucket,
        "visit_filter": visit_filter, "selected_blc_case": blc_case,
        "type_of_notice": type_of_notice, "company": company,
        "sort_field": sort_field, "sort_order": sort_order,
    }

    return render(request, "financehub/executive_visit_schedule_list.html", context)

@financehub_required
def executive_visit_schedule_edit(request, pk):
    obj = ExecutiveVisitScheduling.objects.get(pk=pk)

    if request.method == "POST":
        obj.loanno = request.POST.get("loanno")
        obj.empid = request.POST.get("empid")
        obj.visit_schedule_date = request.POST.get("visit_schedule_date")
        obj.save()

        messages.success(request, "Visit schedule updated successfully")
        return redirect("executive_visit_schedule_list")

    return render(request, "financehub/executive_visit_schedule_edit.html", {
        "obj": obj
    })





@financehub_required
def executive_my_visits(request):
    empid = request.user.username  # employee_id

    visits = ExecutiveVisitScheduling.objects.filter(
        empid=empid
    ).order_by("visit_schedule_date")

    # collect loan numbers
    loan_numbers = list(
        visits.values_list("loanno", flat=True)
    )

    # fetch LCC data
    lcc_map = {
        l.loan_number: l
        for l in Lcc.objects.filter(loan_number__in=loan_numbers)
    }

    # merge data
    data = []
    for v in visits:
        lcc = lcc_map.get(v.loanno)

        data.append({
            "visit": v,                          # IMPORTANT
            "loan_number": v.loanno,
            "visit_date": v.visit_schedule_date,
            "status": v.visit_status,
            "customer_name": lcc.customer_name if lcc else "",
            "vehicle_no": lcc.vehicle_no if lcc else "",
        })

    return render(request, "financehub/executive_my_visits.html", {
        "data": data
    })





from django.utils import timezone

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone

from .models import (
    ExecutiveVisitScheduling,
    CollectionAllocations,   # ✅ REQUIRED
)


from django.utils import timezone
@financehub_required
def executive_visit_response(request, pk):

    visit = get_object_or_404(ExecutiveVisitScheduling, pk=pk)
    next_url = request.POST.get("next") or request.GET.get("next")

    # EXEC CHECK (unchanged logic)
    is_exec = CollectionAllocations.objects.filter(
        employee_id=request.user.username
    ).exists()

    if request.method == "POST":

        # ---------------- RESPONSE (ALWAYS)
        visit.visit_status = request.POST.get("visit_status")

        reason = request.POST.get("not_visited_reason", "").strip()
        if visit.visit_status == "not_visited":
            visit.not_visited_reason = reason
        else:
            visit.not_visited_reason = ""   # clear old reason safely

        # ---------------- SCHEDULING (ONLY IF PROVIDED & NOT EXEC)
        if not is_exec:

            empid = request.POST.get("empid")
            visit_date = request.POST.get("visit_schedule_date")

            # ✅ ONLY UPDATE IF USER CHANGED IT
            if empid:
                visit.empid = empid

            if visit_date:
                visit.visit_schedule_date = visit_date

        visit.save()
        return redirect(next_url or "executive_visit_schedule_list")

    return render(request, "financehub/executive_visit_response.html", {
        "visit": visit,
        "next": next_url,
        "is_exec": is_exec,
    })




from django.db.models import Q
from django.core.paginator import Paginator
from .models import DueNotice

@financehub_required
def due_notice_list(request):

    query = request.GET.get("q", "").strip()
    qs = DueNotice.objects.all().order_by("-created_at")

    # 🔍 SEARCH
    if query:
        qs = qs.filter(
            Q(loan_number__icontains=query) |
            Q(bar_number__icontains=query) |
            Q(vehicle_no__icontains=query) |
            Q(customer_name__icontains=query) |
            Q(company__icontains=query) |
            Q(branch__icontains=query) |
            Q(send_to__icontains=query) |
            Q(type_of_notice__icontains=query) |
            Q(notice_status__icontains=query)
        )

    # ✅ AJAX UPDATE
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            notice_id = data.get("notice_id")
            status = data.get("notice_status")
            delivery_date = data.get("delivery_date")
            return_date = data.get("return_date")
            return_reason = data.get("return_reason", "").strip()

            if not notice_id or not status:
                return JsonResponse({"success": False, "error": "Missing data"})

            update_data = {"notice_status": status}

            if status == DueNotice.NoticeStatus.DELIVERED:
                if not delivery_date:
                    return JsonResponse({"success": False, "error": "Delivery date required"})
                update_data["delivery_date"] = delivery_date
                update_data["return_date"] = None
                update_data["return_reason"] = None

            elif status == DueNotice.NoticeStatus.RETURNED:
                if not return_date:
                    return JsonResponse({"success": False, "error": "Return date required"})
                update_data["return_date"] = return_date
                update_data["return_reason"] = return_reason
                update_data["delivery_date"] = None

            else:  # IN TRANSIT
                update_data["delivery_date"] = None
                update_data["return_date"] = None
                update_data["return_reason"] = None  # ✅ FIXED

            DueNotice.objects.filter(id=notice_id).update(**update_data)

            return JsonResponse({"success": True})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    # 📄 PAGINATION
    paginator = Paginator(qs, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "financehub/due_notice_list.html",
        {
            "page_obj": page_obj,
            "query": query,
            "status_choices": DueNotice.NoticeStatus.choices,
        },
    )



# ======================== DELETE OPTION FOR LCC, C ALLOCATION, REPO, PAID, CLOSED =========================

@financehub_required
def lcc_delete(request):
    lcc = Lcc.objects.all()
    lcc.delete()
    return render(request,'financehub/upload.html')



@financehub_required
def repo_delete(request):
    repo = Repo.objects.all()
    repo.delete()
    return render(request,'financehub/upload.html')

@financehub_required
def cA_delete(request):
    ca = CollectionAllocations.objects.all()
    ca.delete()
    return render(request,'financehub/upload.html')

@financehub_required
def paid_delete(request):
    paid = Paid.objects.all()
    paid.delete()
    return render(request,'financehub/upload.html')

@financehub_required
def closed_delete(request):
    closed = Closed.objects.all()
    closed.delete()
    return render(request,'financehub/upload.html')







@financehub_required
def download_employee_report(request):

    import calendar
    import datetime
    import pandas as pd
    from io import BytesIO
    from django.http import HttpResponse
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    # =====================================================
    # GET DATE RANGE
    # =====================================================
    from_date_param = request.GET.get("from_date", "").strip()
    to_date_param = request.GET.get("to_date", "").strip()

    # =====================================================
    # LOAD EMPLOYEE MASTER
    # =====================================================
    master_qs = EmployeeMaster.objects.all().values(
        "employee_number",
        "employee_name",
        "joined_on",
        "curr_designation",
        "curr_department",
        "curr_location",
        "phone",
        "reporting_to_collections",
        "status",
        "lwd"
    )

    master_df = pd.DataFrame(list(master_qs))

    if master_df.empty:
        return HttpResponse("No Employee Master data found.")

    master_df["status"] = (
        master_df["status"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    master_df["joined_on"] = pd.to_datetime(
        master_df["joined_on"],
        format="%d-%b-%y",
        errors="coerce"
    )

    master_df["lwd"] = pd.to_datetime(
        master_df["lwd"],
        format="%d %b %Y",
        errors="coerce"
    )

    master_df["phone"] = (
        master_df["phone"]
        .astype(str)
        .str.split(".")
        .str[0]
    )

    active_employees = master_df[
        (master_df["status"] != "LEFT")
        &
        (
            master_df["curr_department"]
            .str.upper()
            .isin(["COLLECTION", "SALES/COLLECTION"])
        )
    ]

    active_employees = active_employees[
        active_employees["curr_location"].str.upper()
        != "HEAD OFFICE - RANIGUNJ"
    ]

    # =====================================================
    # LOAD CLU
    # =====================================================
    clu_qs = Clu.objects.values(
        "employee_id",
        "employee_name",
        "visited_on"
    )

    clu_df = pd.DataFrame(list(clu_qs))

    if not clu_df.empty:

        clu_df = clu_df[
            clu_df["employee_id"].notnull()
        ]

        clu_df["visit_datetime"] = pd.to_datetime(
            clu_df["visited_on"],
            format="%b %d,%Y, %I:%M:%S %p",
            errors="coerce"
        )

        clu_df["visit_date"] = (
            clu_df["visit_datetime"]
            .dt.normalize()
        )

    # =====================================================
    # DATE RANGE LOGIC
    # =====================================================
    try:

        # ---------------------------------
        # CUSTOM DATE RANGE
        # ---------------------------------
        if from_date_param and to_date_param:

            start_date = datetime.datetime.strptime(
                from_date_param,
                "%Y-%m-%d"
            ).date()

            end_date = datetime.datetime.strptime(
                to_date_param,
                "%Y-%m-%d"
            ).date()

            if start_date > end_date:
                return HttpResponse(
                    "From Date cannot be greater than To Date"
                )

            if not clu_df.empty:

                clu_df = clu_df[
                    (clu_df["visit_date"] >= pd.Timestamp(start_date))
                    &
                    (clu_df["visit_date"] <= pd.Timestamp(end_date))
                ]

        # ---------------------------------
        # CURRENT MONTH (OLD LOGIC)
        # ---------------------------------
        else:

            if not clu_df.empty:

                max_date = clu_df["visit_date"].max()

                if pd.isna(max_date):
                    max_date = pd.Timestamp(
                        datetime.date.today()
                    )

            else:

                max_date = pd.Timestamp(
                    datetime.date.today()
                )

            year = max_date.year
            month = max_date.month

            start_date = datetime.date(
                year,
                month,
                1
            )

            end_date = datetime.date(
                year,
                month,
                calendar.monthrange(year, month)[1]
            )

    except Exception as e:

        return HttpResponse(
            f"Invalid date range selected. Error: {str(e)}"
        )

    # =====================================================
    # GENERATE DATE LIST
    # =====================================================
    all_days_list = pd.date_range(
        start=start_date,
        end=end_date
    )

    # =====================================================
    # CLEAN CLU
    # =====================================================
    if not clu_df.empty:

        clu_df = clu_df.sort_values(
            by=["visit_datetime"]
        )

        clu_df = clu_df.groupby(
            ["employee_id", "visit_date"],
            as_index=False
        ).first()

    # =====================================================
    # EMPLOYEE x DAY MATRIX
    # =====================================================
    employees = active_employees[
        [
            "employee_number",
            "employee_name",
            "joined_on"
        ]
    ].drop_duplicates()

    all_days = pd.DataFrame(
        [
            (eid, ename, jdate, day)
            for eid, ename, jdate in employees.values
            for day in all_days_list
        ],
        columns=[
            "employee_id",
            "employee_name",
            "joined_on",
            "visit_date"
        ]
    )

    # =====================================================
    # MERGE VISITS
    # =====================================================
    if not clu_df.empty:

        merged = pd.merge(
            all_days,
            clu_df[
                [
                    "employee_id",
                    "visit_date",
                    "visit_datetime"
                ]
            ],
            on=[
                "employee_id",
                "visit_date"
            ],
            how="left"
        )

    else:

        merged = all_days.copy()
        merged["visit_datetime"] = None

    # =====================================================
    # PIVOT
    # =====================================================
    day_df = merged.pivot(
        index=[
            "employee_id",
            "employee_name",
            "joined_on"
        ],
        columns="visit_date",
        values="visit_datetime"
    ).reset_index()

    day_df = day_df[
        [
            "employee_id",
            "employee_name",
            "joined_on"
        ]
        + list(all_days_list)
    ]

    # =====================================================
    # ATTENDANCE LOGIC
    # =====================================================
    for day in all_days_list:

        day_df[day] = day_df.apply(
            lambda row:
            (
                "Not Yet Joined"
                if (
                    pd.notna(row["joined_on"])
                    and row["joined_on"] > day
                )
                else (
                    None
                    if pd.isna(row[day])
                    else row[day]
                )
            ),
            axis=1
        )

    # =====================================================
    # FINAL MERGE
    # =====================================================
    final_df = pd.merge(
        active_employees,
        day_df,
        left_on=[
            "employee_number",
            "employee_name",
            "joined_on"
        ],
        right_on=[
            "employee_id",
            "employee_name",
            "joined_on"
        ],
        how="left"
    )

    cols = (
        list(active_employees.columns)
        + list(all_days_list)
    )

    final_df = final_df[cols]

    final_df["joined_on"] = (
        final_df["joined_on"]
        .dt.strftime("%Y-%m-%d")
    )

    final_df["status"] = final_df["status"].replace(
        ["NULL", None, ""],
        "NONE"
    )

    # =====================================================
    # COLUMN NAMES
    # =====================================================
    final_df.columns = [
        (
            col.strftime("%d-%b (%a)")
            if isinstance(col, pd.Timestamp)
            else col
        )
        for col in final_df.columns
    ]

    # =====================================================
    # EXCEL EXPORT
    # =====================================================
    output = BytesIO()
    final_df.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    day_to_col_index = {
        day: 11 + idx
        for idx, day in enumerate(all_days_list)
    }

    for row_idx in range(2, ws.max_row + 1):

        for day_date, col_idx in day_to_col_index.items():

            cell = ws.cell(
                row=row_idx,
                column=col_idx
            )

            if cell.value == "Not Yet Joined":

                cell.font = Font(color="000000")

            elif cell.value == "Sunday":

                cell.font = Font(color="0000FF")

            elif cell.value == "Absent":

                cell.font = Font(color="000000")

            elif isinstance(
                cell.value,
                datetime.datetime
            ):

                if cell.value.time() > datetime.time(7, 0):

                    cell.font = Font(
                        color="FF0000"
                    )

                else:

                    cell.font = Font(
                        color="000000"
                    )

                cell.number_format = (
                    "MMM DD,YYYY, hh:mm:ss AM/PM"
                )

            elif cell.value is None:

                if day_date.weekday() == 6:

                    cell.value = "Sunday"
                    cell.font = Font(
                        color="0000FF"
                    )

                else:

                    cell.value = "Absent"
                    cell.font = Font(
                        color="000000"
                    )

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # =====================================================
    # FILE NAME
    # =====================================================
    if from_date_param and to_date_param:

        filename = (
            f"Employee_Report_"
            f"{from_date_param}"
            f"_to_"
            f"{to_date_param}.xlsx"
        )

    else:

        filename = (
            f"Employee_Report_"
            f"{start_date.strftime('%Y-%m')}.xlsx"
        )

    response = HttpResponse(
        final_output,
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    return response

@financehub_required
def employee_report_page(request):
    return render(request, "financehub/employee_report.html")




@financehub_required
def employee_monthly_attendance(request):
    """Display monthly attendance with CLU visit timings for selected employee"""

    employee_id = request.GET.get('employee_id', '').strip()
    year = request.GET.get('year', str(datetime.datetime.now().year))
    month = request.GET.get('month', str(datetime.datetime.now().month))

    context = {
        'employee': None,
        'attendance_data': [],
        'year': int(year),
        'month': int(month),
        'months': range(1, 13),
        'years': range(2020, datetime.datetime.now().year + 1),
        'employee_search': employee_id,
        'error': None
    }

    # Get employee details
    if employee_id:
        try:
            employee = EmployeeMaster.objects.get(employee_number=employee_id)
            context['employee'] = employee
        except EmployeeMaster.DoesNotExist:
            context['error'] = f"Employee with ID '{employee_id}' not found"
            return render(request, "financehub/employee_monthly_attendance.html", context)

        # Get CLU data for the employee for selected month
        year_int = int(year)
        month_int = int(month)

        # Create date range for the month
        start_date = datetime.date(year_int, month_int, 1)
        if month_int == 12:
            end_date = datetime.date(year_int + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            end_date = datetime.date(year_int, month_int + 1, 1) - datetime.timedelta(days=1)

        # Get all CLU visits for this employee in the date range
        clu_visits = Clu.objects.filter(
            employee_id=employee_id,
            visited_on__isnull=False
        ).order_by('visited_on')

        # Parse dates and create attendance map
        attendance_map = {}

        for visit in clu_visits:
            try:
                # Parse visited_on date with multiple formats
                visited_value = str(visit.visited_on).strip()
                visited_date = None

                # Try different date formats
                date_formats = [
                    "%Y-%m-%d %H:%M:%S",
                    "%Y-%m-%d",
                    "%d-%b-%Y %H:%M:%S",
                    "%d-%b-%Y %I:%M %p",
                    "%b %d,%Y, %I:%M:%S %p",
                    "%d/%m/%Y",
                ]

                for fmt in date_formats:
                    try:
                        visited_date = datetime.datetime.strptime(visited_value, fmt)
                        break
                    except (ValueError, TypeError):
                        continue

                if visited_date:
                    date_key = visited_date.date()
                    # Only include if within selected month
                    if start_date <= date_key <= end_date:
                        attendance_map[date_key] = {
                            'time': visited_date.time(),
                            'datetime': visited_date,
                            'remarks': getattr(visit, 'remarks', '') or '',
                            'customer_name': getattr(visit, 'customer_name', '') or '',
                            'loan_number': getattr(visit, 'loan_number', '') or '',
                            'status': getattr(visit, 'status', '') or '',
                            'type_of_visit': getattr(visit, 'type_of_visit', '') or ''
                        }
            except Exception as e:
                print(f"Error parsing date: {visit.visited_on}, Error: {e}")
                continue

        # Build attendance data for all days of the month
        attendance_data = []
        current_date = start_date
        while current_date <= end_date:
            visit_info = attendance_map.get(current_date)

            day_data = {
                'date': current_date,
                'day_name': current_date.strftime('%A'),
                'day_number': current_date.day,
                'is_weekend': current_date.weekday() >= 5,  # Saturday=5, Sunday=6
                'has_visit': visit_info is not None,
                'visit_time': visit_info['time'].strftime('%I:%M %p') if visit_info and visit_info['time'] else None,
                'visit_datetime': visit_info['datetime'] if visit_info else None,
                'is_late': visit_info and visit_info['time'] and visit_info['time'] > datetime.time(7, 0) if visit_info else False,
                'remarks': visit_info['remarks'] if visit_info else '',
                'customer_name': visit_info['customer_name'] if visit_info else '',
                'loan_number': visit_info['loan_number'] if visit_info else '',
                'status': visit_info['status'] if visit_info else '',
                'type_of_visit': visit_info['type_of_visit'] if visit_info else ''
            }
            attendance_data.append(day_data)
            current_date += datetime.timedelta(days=1)

        # Calculate statistics
        total_days = len(attendance_data)
        working_days = sum(1 for d in attendance_data if not d['is_weekend'])
        holidays = sum(1 for d in attendance_data if d['is_weekend'])
        days_present = sum(1 for d in attendance_data if d['has_visit'])
        days_absent = working_days - days_present
        late_visits = sum(1 for d in attendance_data if d['is_late'])
        on_time_visits = days_present - late_visits
        attendance_percentage = (days_present / working_days * 100) if working_days > 0 else 0

        context.update({
            'attendance_data': attendance_data,
            'total_days': total_days,
            'working_days': working_days,
            'holidays': holidays,
            'days_present': days_present,
            'days_absent': days_absent,
            'late_visits': late_visits,
            'on_time_visits': on_time_visits,
            'attendance_percentage': round(attendance_percentage, 2),
            'month_name': start_date.strftime('%B'),
            'year': year_int,
            'month': month_int
        })

    return render(request, "financehub/employee_monthly_attendance.html", context)
