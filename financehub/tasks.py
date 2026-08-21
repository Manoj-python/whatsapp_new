# financehub/tasks.py

from celery import shared_task
import openpyxl
import pandas as pd
import os
from django.apps import apps
from django.core.cache import cache

from .utils import clean_header, get_model_by_type
from .models import UploadHistory, LoanStatusCache

BULK_BATCH_SIZE = 2000
PANDAS_CHUNK_SIZE = 5000


# ---------------------------------------------------------
# SAFE DATE PARSER
# ---------------------------------------------------------
def parse_datetime_safe(value):
    if value in (None, "", "nan", "NaT"):
        return None

    try:
        if isinstance(value, str) and (("AM" in value.upper()) or ("PM" in value.upper())):
            try:
                return pd.to_datetime(value, format="%b %d,%Y, %I:%M:%S %p").to_pydatetime()
            except:
                pass

        dt = pd.to_datetime(value, errors="coerce")
        if pd.isna(dt):
            return None

        return dt.to_pydatetime()

    except Exception:
        return None

from django.db import transaction
from django.utils import timezone
from .models import EmployeeMaster, CollectionAllocations

def handle_employee_reassignment(df, model_name):
    """
    Check for previous_employee_id column and perform reassignment
    """
    try:
        # Only for CollectionAllocations
        if model_name != "CollectionAllocations":
            return None
        
        # Check if required columns exist
        if 'previous_employee_id' not in df.columns or 'employee_id' not in df.columns:
            return None
        
        # Get unique employee reassignments with executive_name if available
        if 'executive_name' in df.columns:
            reassign_df = df[['previous_employee_id', 'employee_id', 'executive_name']].drop_duplicates()
            reassign_df = reassign_df.dropna(subset=['previous_employee_id', 'employee_id'])
            print(f"📋 Found executive_name column in file")
        else:
            reassign_df = df[['previous_employee_id', 'employee_id']].drop_duplicates()
            reassign_df = reassign_df.dropna()
            print(f"📋 No executive_name column found in file")
        
        if reassign_df.empty:
            return None
        
        total_reassigned = 0
        reassignment_results = []
        
        with transaction.atomic():
            for _, row in reassign_df.iterrows():
                old_id = str(row['previous_employee_id']).strip()
                new_id = str(row['employee_id']).strip()
                
                if not old_id or not new_id or old_id == new_id:
                    continue
                
                # 👇 PRIORITY: Use executive_name from file if available
                if 'executive_name' in row and pd.notna(row.get('executive_name')):
                    new_name = str(row['executive_name']).strip()
                    print(f"✅ Using executive_name from file: {new_name}")
                else:
                    # Fallback: Get from EmployeeMaster
                    new_employee = EmployeeMaster.objects.filter(
                        employee_number=new_id
                    ).first()
                    new_name = new_employee.employee_name if new_employee else f"Emp_{new_id}"
                    print(f"✅ Using executive_name from EmployeeMaster: {new_name}")
                
                # Find cases assigned to old employee
                cases = CollectionAllocations.objects.filter(employee_id=old_id)
                case_count = cases.count()
                
                print(f"🔄 Found {case_count} cases for employee {old_id}")
                
                if case_count > 0:
                    # Show sample loan numbers
                    sample_cases = cases[:3]
                    sample_loans = [c.loan_number for c in sample_cases]
                    print(f"📋 Sample cases: {', '.join(sample_loans)}")
                    
                    # Update each case
                    updated_cases = []
                    for case in cases:
                        case.previous_employee_id = old_id
                        case.employee_id = new_id
                        case.executive_name = new_name
                        case.reassigned_at = timezone.now()
                        updated_cases.append(case)
                    
                    # Bulk update
                    CollectionAllocations.objects.bulk_update(
                        updated_cases,
                        ['employee_id', 'executive_name', 'previous_employee_id', 'reassigned_at']
                    )
                    
                    total_reassigned += case_count
                    reassignment_results.append({
                        'old_id': old_id,
                        'new_id': new_id,
                        'count': case_count,
                        'status': 'Reassigned'
                    })
                    
                    print(f"✅ Successfully reassigned {case_count} cases from {old_id} to {new_id}")
        
        return {
            'total_reassigned': total_reassigned,
            'details': reassignment_results
        }
        
    except Exception as e:
        print(f"❌ Reassignment error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


# ---------------------------------------------------------
# UNIVERSAL PROCESSOR
# ---------------------------------------------------------
@shared_task(bind=True)
def process_universal_file(self, upload_id, tmp_path, ext, file_type):

    upload = UploadHistory.objects.get(id=upload_id)
    reassignment_result = None

    try:
        upload.status = "processing"
        upload.save(update_fields=["status"])

        Model = get_model_by_type(file_type)
        if not Model:
            upload.status = "error"
            upload.error_message = "Invalid file type"
            upload.save()
            return

        model_fields = {f.name for f in Model._meta.fields}
        processed_rows = 0
        model_name = Model.__name__

        # ================= UNIQUE LOGIC =================
        unique_field = None

        if Model.__name__ in ["Lcc"]:
            unique_field = "loan_number"

        elif Model.__name__ in ["Clu", "Dialer", "DueNotice", "EmployeeMaster", "Paid", "CollectionAllocations","OpenRepo"]:
            unique_field = None

        else:
            for field in ["loan_number", "agreement_number", "employee_number", "ticket_id","loan_no"]:
                if field in model_fields:
                    unique_field = field
                    break

        existing_values = set()
        if unique_field:
            existing_values = set(
                Model.objects.values_list(unique_field, flat=True)
            )

        # CLU duplicate prevention
        clu_existing = set()

        if Model.__name__ == "Clu":
            clu_existing = set(
                Model.objects.values_list(
                    "employee_id",
                    "visited_on"
                )
            )

        # =====================================================
        # ====================== CSV ===========================
        # =====================================================
        if ext == "csv":

            def process_chunk(chunk):
                nonlocal processed_rows, existing_values
                # ❌ REMOVED: nonlocal reassignment_result

                # ✅ CRITICAL FIX: Pass model_name to clean_header
                headers = [clean_header(h, model_name) for h in chunk.columns]
                chunk.columns = headers
                header_map = {h: h for h in headers if h in model_fields}

                upload.total_rows += len(chunk)
                upload.save(update_fields=["total_rows"])

                batch = []

                for row in chunk.to_dict("records"):
                    cleaned = {}

                    for col, val in row.items():
                        if col not in header_map:
                            continue

                        field = Model._meta.get_field(col)

                        if isinstance(val, str):
                            val = val.replace('\xa0', ' ').strip()

                        field_type = field.get_internal_type()

                        # visited_on special
                        if col == "visited_on":
                            dt = parse_datetime_safe(val)
                            cleaned[col] = dt if dt else None

                        elif field_type in ["DateField", "DateTimeField"]:
                            cleaned[col] = parse_datetime_safe(val)

                        elif field_type == "IntegerField":
                            try:
                                cleaned[col] = int(val) if val else None
                            except:
                                cleaned[col] = None

                        elif field_type == "DecimalField":
                            try:
                                cleaned[col] = float(val) if val else None
                            except:
                                cleaned[col] = None

                        else:
                            cleaned[col] = val if val is not None else ""

                    # DUPLICATE CHECK
                    if unique_field:
                        key = cleaned.get(unique_field)
                        if not key or key in existing_values:
                            continue
                        existing_values.add(key)

                    # CLU duplicate check
                    if Model.__name__ == "Clu":
                        visit_key = (
                            str(cleaned.get("employee_id", "")).strip(),
                            str(cleaned.get("visited_on", "")).strip(),
                        )

                        if visit_key in clu_existing:
                            continue

                        clu_existing.add(visit_key)

                    # 👉 SKIP REASSIGNMENT-ONLY ROWS (NO loan_number)
                    if model_name == "CollectionAllocations":
                        if not cleaned.get('loan_number'):
                            continue

                    batch.append(Model(**cleaned))
                    processed_rows += 1

                    if len(batch) >= BULK_BATCH_SIZE:
                        Model.objects.bulk_create(batch, ignore_conflicts=True)
                        batch = []

                if batch:
                    Model.objects.bulk_create(batch, ignore_conflicts=True)

                upload.processed_rows = processed_rows
                upload.save(update_fields=["processed_rows"])

            reader = pd.read_csv(tmp_path, dtype=str, chunksize=PANDAS_CHUNK_SIZE)
            all_rows_data = []

            for chunk in reader:
                if model_name == "CollectionAllocations":
                    chunk_dict = chunk.to_dict('records')
                    all_rows_data.extend(chunk_dict)
                process_chunk(chunk)

            # 👉 PERFORM REASSIGNMENT AFTER ALL DATA IS INSERTED
            if model_name == "CollectionAllocations" and all_rows_data:
                df = pd.DataFrame(all_rows_data)
                reassignment_result = handle_employee_reassignment(df, model_name)

        # =====================================================
        # ==================== EXCEL ===========================
        # =====================================================
        elif ext in ("xlsx", "xls"):

            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb.active

            raw_headers = next(ws.iter_rows(values_only=True))
            headers = [clean_header(h, model_name) for h in raw_headers]

            header_map = {h: h for h in headers if h in model_fields}

            upload.total_rows = max(ws.max_row - 1, 0)
            upload.save(update_fields=["total_rows"])

            batch = []
            all_rows_data = []

            for row in ws.iter_rows(min_row=2, values_only=True):

                if not row or all(v in (None, "", " ") for v in row):
                    continue

                row_dict = dict(zip(headers, row))

                if model_name == "CollectionAllocations":
                    all_rows_data.append(row_dict)

                cleaned = {}

                for col, val in row_dict.items():
                    if col not in header_map:
                        continue

                    field = Model._meta.get_field(col)

                    if isinstance(val, str):
                        val = val.strip()

                    field_type = field.get_internal_type()

                    if col == "visited_on":
                        cleaned[col] = val.strip() if val else None

                    elif field_type in ["DateField", "DateTimeField"]:
                        cleaned[col] = parse_datetime_safe(val)

                    elif field_type == "IntegerField":
                        try:
                            cleaned[col] = int(val) if val else None
                        except:
                            cleaned[col] = None

                    elif field_type == "DecimalField":
                        try:
                            cleaned[col] = float(val) if val else None
                        except:
                            cleaned[col] = None

                    else:
                        cleaned[col] = val if val is not None else ""

                if unique_field:
                    key = cleaned.get(unique_field)
                    if not key or key in existing_values:
                        continue
                    existing_values.add(key)

                # CLU duplicate check
                if Model.__name__ == "Clu":
                    visit_key = (
                        str(cleaned.get("employee_id", "")).strip(),
                        str(cleaned.get("visited_on", "")).strip(),
                    )

                    if visit_key in clu_existing:
                        continue

                    clu_existing.add(visit_key)

                # 👉 SKIP REASSIGNMENT-ONLY ROWS (NO loan_number)
                if model_name == "CollectionAllocations":
                    if not cleaned.get('loan_number'):
                        continue

                batch.append(Model(**cleaned))
                processed_rows += 1

                if len(batch) >= BULK_BATCH_SIZE:
                    Model.objects.bulk_create(batch, ignore_conflicts=True)
                    batch = []

            if batch:
                Model.objects.bulk_create(batch, ignore_conflicts=True)

            # 👉 PERFORM REASSIGNMENT AFTER INSERTING DATA
            if model_name == "CollectionAllocations" and all_rows_data:
                df = pd.DataFrame(all_rows_data)
                reassignment_result = handle_employee_reassignment(df, model_name)

        # FINAL UPDATE
        upload.processed_rows = processed_rows
        upload.status = "completed"
        upload.save()

        if Model.__name__ == "Lcc":
            cache.delete('dropdowns_final_v3')
            deleted_count = LoanStatusCache.objects.all().delete()
            print(f"Cleared LoanStatusCache after LCC upload")

        # Log reassignment result
        if reassignment_result and reassignment_result.get('total_reassigned', 0) > 0:
            print(f"✅ Reassigned {reassignment_result['total_reassigned']} cases")
            for detail in reassignment_result.get('details', []):
                print(f"   {detail['old_id']} → {detail['new_id']}: {detail['count']} cases")

    except Exception as e:
        upload.status = "error"
        upload.error_message = str(e)
        upload.save()
        import traceback
        traceback.print_exc()

    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except:
            pass
