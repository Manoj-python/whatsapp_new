import pandas as pd
import io
import json
import re
import uuid
import requests
import time
import traceback
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db.models import Max, Count, Q, F
from django.conf import settings
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.utils import timezone

from .models import *
from .utils import *
from .tasks import *
from .forms import UploadForm2
from django.contrib.auth import authenticate
from django.contrib import messages
from django.core.paginator import Paginator
from .consumers import *
from django.http import StreamingHttpResponse, HttpResponseForbidden, Http404
from django.shortcuts import get_object_or_404
import pytz
import os
import tempfile
from django.conf import settings
from .models import *
from financehub.models import *
from django.shortcuts import render
from django.db.models import Q,Value
from django.core.paginator import Paginator
from django.db.models.functions import Replace
import unicodedata
import uuid
import requests
import time
import traceback
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db.models import Max, Count, Q, F
from django.conf import settings
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.utils import timezone
from django.contrib.auth import authenticate, login as auth_login
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required

from .models import *
from .utils import *
from .tasks import *
from .forms import UploadForm2
from .consumers import *

import pytz
import os
import tempfile
from financehub.models import *
from django.db.models.functions import Replace
import unicodedata

# Import the unified app config from adminpanel
from adminpanel.views import APP_CONFIG, get_agent_from_user

# ============================================
# HELPER: get models for selected app
# ============================================
def get_models_for_app(request):
    """Return (case_model, contact_model, log_model, channel_group) for the app selected via ?app="""
    app_key = request.GET.get('app', 'psf')
    if app_key not in APP_CONFIG:
        app_key = 'psf'
    cfg = APP_CONFIG[app_key]
    return cfg['case_model'], cfg['contact_model'], cfg['log_model'], cfg['channel_group']

def get_case_model_for_app(request):
    """Convenience: only case model"""
    return get_models_for_app(request)[0]

def get_contact_model_for_app(request):
    """Convenience: only contact model"""
    return get_models_for_app(request)[1]

# ============================================
# AUTHENTICATION (unchanged)
# ============================================
def messaging2_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(username=username, password=password)
        if user:
            auth_login(request, user)
            agent, created = Agent.objects.get_or_create(
                user=user,
                defaults={
                    'agent_id': f"AGT-{user.id}",
                    'name': user.get_full_name() or user.username,
                    'email': user.email or f"{user.username}@example.com",
                    'role': 'ADMIN' if user.is_superuser else 'AGENT'
                }
            )
            request.session["messaging2_user"] = user.id

    return render(request, "messaging2/login.html")

def messaging2_logout(request):
    request.session.pop("messaging2_user", None)
    return redirect("messaging2_login")

def messaging2_required(view_func):
    def wrapper(request, *args, **kwargs):
        # First check Django auth
        if request.user.is_authenticated:
            return view_func(request, *args, **kwargs)
        # Then check custom session key (legacy)
        if request.session.get("messaging2_user"):
            return view_func(request, *args, **kwargs)
        return redirect(settings.LOGIN_URL)
    return wrapper
def serialize_log(m):
    return {
        "id": m.id,
        "message_id": m.message_id,
        "mobile": m.mobile,
        "sent_text_message": m.sent_text_message,
        "message_type": m.message_type,
        "content_type": m.content_type,
        "media_file": m.media_file.url if m.media_file else "",
        "sent_at": m.sent_at.isoformat() if m.sent_at else "",
        "status": m.status,
    }

def ws_group2(mobile: str) -> str:
    if not mobile:
        return ""
    return re.sub(r"\D", "", str(mobile))

def send_whatsapp_text2(to_number, text_body):
    url = f"https://graph.facebook.com/v22.0/{settings.WHATSAPP2_PHONE_NUMBER_ID}/messages"
    headers = {
        "Authorization": f"Bearer {settings.WHATSAPP2_ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    payload = {
        "messaging_product": "whatsapp",
        "to": to_number,
        "type": "text",
        "text": {"body": text_body[:4096]},
    }
    resp = requests.post(url, headers=headers, json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json()

def broadcast_delivery2(mobile, message_id, status):
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync
    channel_layer = get_channel_layer()
    status = (status or "").lower()
    if status == "sent":
        norm = "Sent"
    elif status == "delivered":
        norm = "Delivered"
    elif status == "read":
        norm = "Read"
    else:
        norm = "Failed"
    ChatContact2.objects.filter(mobile=mobile).update(last_status=norm)
    gm = ws_group2(mobile)
    if gm:
        async_to_sync(channel_layer.group_send)(
            f"chat2_{gm}",
            {"type": "delivery.update", "message_id": message_id, "status": norm, "mobile": mobile}
        )
    async_to_sync(channel_layer.group_send)(
        "delivery_group2",
        {"type": "delivery.update", "message_id": message_id, "status": norm, "mobile": mobile}
    )
    async_to_sync(channel_layer.group_send)(
        "global_contacts2",
        {"type": "contact.update", "contact": {"mobile": mobile, "last_status": norm}}
    )

# -----------------------------------------------------
# Bulk Upload (PSF only - unchanged)
# -----------------------------------------------------
@messaging2_required
def upload_and_send2(request):
    if request.method == "POST":
        form = UploadForm2(request.POST, request.FILES)
        if form.is_valid():
            choice = form.cleaned_data["template_choice"]
            excel_file = request.FILES["excel_file"]
            unique_name = f"{uuid.uuid4().hex}_{excel_file.name}"
            s3_key = f"uploads2/{unique_name}"
            default_storage.save(s3_key, excel_file)
            with default_storage.open(s3_key, "rb") as f:
                data = f.read()
            df = pd.read_excel(io.BytesIO(data), dtype=str).fillna("")
            job_id = str(uuid.uuid4())
            BulkJob2.objects.create(
                job_id=job_id,
                template_name=choice,
                total_customers=len(df),
                excel_file=s3_key,
                status="Pending",
            )
            process_bulk_whatsapp2.apply_async(args=(s3_key, choice, job_id), queue="messaging2")
            return redirect("job_status2", job_id=job_id)
    else:
        form = UploadForm2()
    return render(request, "messaging2/index.html", {"form": form})

def job_status2(request, job_id):
    job = get_object_or_404(BulkJob2, job_id=job_id)
    progress = 0
    if job.total_customers > 0:
        progress = round((job.sent_count / job.total_customers) * 100, 2)
    return render(request, "messaging2/job_status.html", {"job": job, "progress": progress})

def download_success_report2(request, job_id):
    job = get_object_or_404(BulkJob2, job_id=job_id)
    if job.success_report:
        return redirect(default_storage.url(job.success_report.name))
    raise Http404("Success report not found.")

def download_failed_report2(request, job_id):
    job = get_object_or_404(BulkJob2, job_id=job_id)
    if job.failed_report:
        return redirect(default_storage.url(job.failed_report.name))
    raise Http404("Failed report not found.")

# -----------------------------------------------------
# CHAT DASHBOARD (PSF only)
# -----------------------------------------------------
def chat_dashboard2(request):
    agent = get_agent_from_user(request.user)
    mobiles = (SmsWhatsAppLog2.objects.values("mobile").annotate(last_sent=Max("sent_at")).order_by("-last_sent"))
    seen = set()
    mobile_list = []
    for m in mobiles:
        normalized = format_mobile2(str(m["mobile"]))
        if normalized not in seen:
            seen.add(normalized)
            mobile_list.append({"mobile": normalized})
    return render(request, "messaging2/chat.html", {
        "mobile_list": mobile_list,
        "user_name": request.user.username,
        "MEDIA_URL": settings.MEDIA_URL,
        "agent": agent,
        "user": request.user,
    })

def chat_messages_api2(request, mobile):
    mobile = format_mobile2(mobile)
    page = int(request.GET.get("page", 1))
    size = 500
    qs = SmsWhatsAppLog2.objects.filter(mobile=mobile).order_by("-sent_at")
    paginator = Paginator(qs, size)
    try:
        pg = paginator.page(page)
    except:
        return JsonResponse({"messages": [], "has_more": False})
    result = list(pg.object_list)[::-1]
    def to_json(m):
        media_url = ""
        if m.media_file:
            try:
                media_url = default_storage.url(m.media_file.name)
            except:
                media_url = getattr(m.media_file, "url", "")
        return {
            "id": m.id,
            "mobile": m.mobile,
            "sent_text_message": m.sent_text_message or "",
            "message_type": m.message_type,
            "sent_at": timezone.localtime(m.sent_at).isoformat(),
            "message_id": m.message_id,
            "content_type": m.content_type or "text",
            "media_file": media_url,
            "status": m.status or "",
            "sender_name": m.customer_name or "",
        }
    return JsonResponse({"messages": [to_json(m) for m in result], "has_more": pg.has_next()})

def contacts_api2(request):
    q = request.GET.get("q", "").strip()
    qs = (SmsWhatsAppLog2.objects.values("mobile")
          .annotate(last_time=Max("sent_at"),
                    unread=Count("id", filter=Q(message_type="Received", status="Unread")))
          .order_by("-last_time"))
    if q:
        digits = re.sub(r"\D", "", q)
        if digits:
            qs = qs.filter(mobile__icontains=digits)
        else:
            mobiles_matching = SmsWhatsAppLog2.objects.filter(sent_text_message__icontains=q).values_list("mobile", flat=True).distinct()
            qs = qs.filter(mobile__in=list(mobiles_matching))
    result = [{
        "mobile": format_mobile2(item["mobile"]),
        "last_time": item["last_time"].isoformat() if item["last_time"] else "",
        "unread": item["unread"],
    } for item in qs]
    return JsonResponse({"contacts": result})


import pandas as pd
import io
import json
import re
import uuid
import requests
import time
import traceback
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db.models import Max, Count, Q,F
from django.conf import settings
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.utils import timezone
from django.contrib.auth import authenticate
from django.contrib import messages
from django.core.paginator import Paginator


# =============================================
# HELPER FUNCTIONS
# =============================================


def download_whatsapp_media2(media_id):
    """Download media from WhatsApp"""
    try:
        access_token = settings.WHATSAPP2_ACCESS_TOKEN
        headers = {"Authorization": f"Bearer {access_token}"}

        # Use v22.0
        meta_url = f"https://graph.facebook.com/v22.0/{media_id}"
        meta_resp = requests.get(meta_url, headers=headers, timeout=30)
        meta_resp.raise_for_status()
        meta = meta_resp.json()

        file_url = meta.get("url")
        mime = meta.get("mime_type", "")
        ext = mime.split("/")[-1] if "/" in mime else "bin"

        file_resp = requests.get(file_url, headers=headers, timeout=30)
        file_resp.raise_for_status()

        filename = f"WHATSAPP2_{media_id}.{ext}"
        return filename, file_resp.content

    except Exception as e:
        # print(f"Media download error: {e}")
        return None



# =============================================
# CHAT DASHBOARD & MESSAGES API
# =============================================
from django.core.cache import cache

def chat_messages_api2(request, mobile):
    mobile = format_mobile2(mobile)
    page = int(request.GET.get("page", 1))
    size = 500

    qs = SmsWhatsAppLog2.objects.filter(mobile=mobile).order_by("-sent_at")
    paginator = Paginator(qs, size)

    try:
        pg = paginator.page(page)
    except:
        return JsonResponse({"messages": [], "has_more": False})

    result = list(pg.object_list)[::-1]

    def to_json(m):
        media_url = ""
        if m.media_file:
            try:
                media_url = default_storage.url(m.media_file.name)
            except:
                media_url = getattr(m.media_file, "url", "")
        return {
            "id": m.id,
            "mobile": m.mobile,
            "sent_text_message": m.sent_text_message or "",
            "message_type": m.message_type,
            "sent_at": m.sent_at.isoformat() if m.sent_at else "",
            "message_id": m.message_id,
            "content_type": m.content_type or "text",
            "media_file": media_url,
            "status": m.status or "",
            "sender_name": m.customer_name or "",
        }

    return JsonResponse({
        "messages": [to_json(m) for m in result],
        "has_more": pg.has_next()
    })


@csrf_exempt
def send_reply_api2(request):
    """
    Send reply with media support - SAVE FIRST to prevent NO DB MATCH
    """
    try:
        if request.method != "POST":
            return HttpResponseBadRequest("POST required")

        # Parse request
        if "multipart/form-data" in request.META.get("CONTENT_TYPE", ""):
            mobile = request.POST.get("mobile", "").strip()
            text = request.POST.get("text", "").strip()
            media_file = request.FILES.get("media")
        else:
            payload = json.loads(request.body.decode("utf-8") or "{}")
            mobile = payload.get("mobile", "").strip()
            text = payload.get("text", "").strip()
            media_file = None

        if not mobile:
            return HttpResponseBadRequest("mobile required")

        mobile = format_mobile2(mobile)

        # File size validation
        if media_file:
            file_size_mb = media_file.size / (1024 * 1024)
            file_name = media_file.name.lower()

            if file_name.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
                max_size = 5
            elif file_name.endswith(('.mp4', '.mov', '.avi', '.mkv', '.3gp')):
                max_size = 16
            else:
                max_size = 100

            if file_size_mb > max_size:
                return JsonResponse({
                    "error": f"File too large. Max {max_size}MB. Yours: {file_size_mb:.2f}MB"
                }, status=400)

        # Get agent name
        agent_name = None
        if request.session.get("messaging2_user"):
            from django.contrib.auth.models import User
            u = User.objects.filter(id=request.session["messaging2_user"]).first()
            if u:
                agent_name = u.username

        # =============================================
        # STEP 1: CREATE DATABASE RECORD FIRST (with temp ID)
        # =============================================
        temp_id = str(uuid.uuid4())

        log = SmsWhatsAppLog2.objects.create(
            customer_name=agent_name or "",
            mobile=mobile,
            sent_text_message=text or "",
            status="Sending",  # Status while sending
            message_id=temp_id,
            message_type="Sent",
            content_type="text",  # Will update later
        )
        clear_chat_cache2(mobile)


        # print(f"📝 [STEP 1] Saved pending message with temp ID: {temp_id}")

        # =============================================
        # STEP 2: SEND TO WHATSAPP
        # =============================================
        msg_id = ""
        content_type_val = "text"
        media_url = ""
        saved_path = None

        try:
            if media_file:
                # Determine media type
                file_name = media_file.name.lower()
                original_filename = media_file.name
                if file_name.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
                    WHATSAPP2_media_type = "image"
                    content_type_val = "image"
                elif file_name.endswith(('.mp4', '.mov', '.avi', '.mkv')):
                    WHATSAPP2_media_type = "video"
                    content_type_val = "video"
                elif file_name.endswith(('.mp3', '.wav', '.ogg', '.m4a')):
                    WHATSAPP2_media_type = "audio"
                    content_type_val = "audio"
                else:
                    WHATSAPP2_media_type = "document"
                    content_type_val = "document"

                # Update content type in database
                SmsWhatsAppLog2.objects.filter(id=log.id).update(content_type=content_type_val)

                # print(f"📤 Uploading {WHATSAPP2_media_type}...")

                # Upload to WhatsApp
                upload_resp = upload_whatsapp_media2(media_file)
                media_id = upload_resp.get("id")

                if media_id:
                    send_resp = send_whatsapp_media2(
                        to_number=mobile,
                        media_id=media_id,
                        media_type=WHATSAPP2_media_type,
                        caption=text if text else "",
                        filename=original_filename
                    )
                    msg_id = send_resp.get("messages", [{}])[0].get("id", "")

                    # Save media file to storage
                    media_file.seek(0)
                    saved_path = default_storage.save(
                        f"chat2_media/{media_file.name}",
                        ContentFile(media_file.read())
                    )
                    media_url = default_storage.url(saved_path)

            elif text:
                send_resp = send_whatsapp_text2(mobile, text)
                msg_id = send_resp.get("messages", [{}])[0].get("id", "")

            # print(f"📨 [STEP 2] WhatsApp returned ID: {msg_id}")

            # =============================================
            # STEP 3: UPDATE RECORD WITH REAL WHATSAPP ID
            # =============================================
            if msg_id:
                # Update the existing record
                update_data = {
                    'message_id': msg_id,
                    'status': 'Sent'
                }
                if saved_path:
                    update_data['media_file'] = saved_path

                SmsWhatsAppLog2.objects.filter(id=log.id).update(**update_data)
                log.refresh_from_db()
                # print(f"✅ [STEP 3] Updated message from {temp_id} to {msg_id}")
            else:
                SmsWhatsAppLog2.objects.filter(id=log.id).update(
                    status="Failed",
                    error_message="No message ID returned from WhatsApp"
                )
                # print(f"❌ No WhatsApp ID returned")

        except Exception as e:
            SmsWhatsAppLog2.objects.filter(id=log.id).update(
                status="Failed",
                error_message=str(e)
            )
            # print(f"❌ Send failed: {e}")
            return JsonResponse({"error": f"Send failed: {str(e)}"}, status=500)

        # Update contact
        ChatContact2.objects.update_or_create(
            mobile=mobile,
            defaults={
                "last_msg": text or f"[{content_type_val.title()}]",
                "last_time": timezone.now(),
                "last_type": "Sent",
                "last_status": "Sent" if msg_id else "Failed",
                "unread": 0
            }
        )

        # WebSocket broadcast
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            "global_contacts2",
            {
                "type": "contact.update",
                "contact": {
                    "mobile": mobile,
                    "last_msg": text or f"[{content_type_val.title()}]",
                    "last_time": timezone.now().isoformat(),
                    "last_type": "Sent",
                    "last_status": "Sent" if msg_id else "Failed",
                    "unread": 0
                }
            }
        )
        gm = re.sub(r"\D", "", mobile)

        if gm:
            async_to_sync(channel_layer.group_send)(
                f"chat2_{gm}",
                {
                    "type": "new_message",
                    "message": {
                        "id": log.id,
                        "mobile": mobile,
                        "sent_text_message": log.sent_text_message,
                        "content_type": content_type_val,
                        "media_file": media_url,
                        "sent_at": log.sent_at.isoformat(),
                        "message_type": "Sent",
                        "message_id": log.message_id,  # Now has real ID!
                        "status": log.status,
                        "sender_name": agent_name or "",
                    }
                }
            )

        return JsonResponse({
            "status": "ok",
            "message_id": msg_id,
            "sender_name": agent_name or "",
            "content_type": content_type_val,
            "media_url": media_url
        })

    except Exception as e:
        # print(f"Send reply error: {e}")
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=500)






# =============================================
# WHATSAPP WEBHOOK - COMPLETE FIXED VERSION
# =============================================
@csrf_exempt
def whatsapp_webhook2(request):
    # print("🔥 WEBHOOK HIT")

    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync
    from django.db import transaction
    import json

    channel_layer = get_channel_layer()

    # ---------- VERIFY TOKEN (GET) ----------
    if request.method == "GET":
        mode = request.GET.get("hub.mode")
        token = request.GET.get("hub.verify_token")
        challenge = request.GET.get("hub.challenge")

        if mode == "subscribe" and token == settings.WHATSAPP2_VERIFY_TOKEN:
            return HttpResponse(challenge, status=200)
        return HttpResponseBadRequest("Invalid verification.")

    # ---------- INCOMING WEBHOOK (POST) ----------
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8") or "{}")
            entries = data.get("entry", [])

            for entry in entries:
                for change in entry.get("changes", []):
                    value = change.get("value", {})

                    # ======================================
                    # PROCESS INCOMING MESSAGES
                    # ======================================
                    for msg in value.get("messages", []):
                        msg_id = msg.get("id")
                        mobile = format_mobile2(msg.get("from", ""))

                        if msg_id and SmsWhatsAppLog2.objects.filter(message_id=msg_id).exists():
                            # print(f"⏭️ Duplicate message: {msg_id}")
                            continue

                        msg_type = msg.get("type", "text")
                        text_body = ""
                        content_type = "text"
                        media_file_data = None
                        button_response = ""


                        if msg_type == "text":
                            text_body = msg.get("text", {}).get("body", "").strip()

                            quick_reply_values = [
                                "Interested",
                                "Not Interested",
                                "Call Now"
                            ]

                            if text_body in quick_reply_values:
                                content_type = "interactive"
                                
                                context_id = None
                                if msg.get("context"):
                                    context_id = msg.get("context", {}).get("id")

                                button_response = json.dumps({
                                    "type": "quick_reply_text",
                                    "button_title": text_body,
                                    "button_text": text_body,
                                    "context_message_id": context_id,
                                    "source": "text_quick_reply",
                                    "timestamp": timezone.now().isoformat()
                                })

                                text_body = f"[Button Click] {text_body}"
                                mark_button_clicked(mobile)

                                print(f"🔘 Quick Reply captured via TEXT: {text_body}")

                            else:
                                print(f"📝 Regular text from {mobile}: {text_body}")

                        # ======================================
                        # TEMPLATE QUICK REPLY BUTTONS
                        # ======================================
                        elif msg_type == "button":

                            button = msg.get("button", {})

                            button_text = button.get("text", "")
                            button_payload = button.get("payload", "")

                            content_type = "button"

                            button_response = json.dumps({
                                "type": "template_quick_reply",
                                "button_text": button_text,
                                "button_payload": button_payload,
                                "source": "button",
                                "timestamp": timezone.now().isoformat()
                            })

                            text_body = f"[Button Click] {button_text}"
                            mark_button_clicked(mobile)
 
                            print(
                                f"🔘 Template Button Clicked: "
                                f"text={button_text}, payload={button_payload}"
                            )

                            


                            # print(f"📝 Text from {mobile}: {text_body[:50]}")

                        elif msg_type == "interactive":

                            interactive = msg.get("interactive", {})
                            content_type = "interactive"

                            interactive_type = interactive.get("type")

                            if interactive_type == "button_reply":

                                button_reply = interactive.get("button_reply", {})

                                button_id = button_reply.get("id", "")
                                button_title = button_reply.get("title", "")

                                button_response = json.dumps({
                                    "type": "button_click",
                                    "button_id": button_id,
                                    "button_title": button_title,
                                    "source": "interactive_button_reply",
                                    "timestamp": timezone.now().isoformat()
                                })
                                

                                text_body = f"[Button Click] {button_title}"
                                mark_button_clicked(mobile)

                            elif interactive_type == "list_reply":

                                list_reply = interactive.get("list_reply", {})

                                list_id = list_reply.get("id", "")
                                list_title = list_reply.get("title", "")

                                button_response = json.dumps({
                                    "type": "list_selection",
                                    "list_id": list_id,
                                    "list_title": list_title,
                                    "source": "interactive_list",
                                    "timestamp": timezone.now().isoformat()
                                })

                                text_body = f"[List Selection] {list_title}"
                                mark_button_clicked(mobile)

                            elif interactive_type == "cta_url":

                                cta = interactive.get("cta_url", {})

                                button_title = cta.get("title", "")

                                button_response = json.dumps({
                                    "type": "cta_click",
                                    "button_title": button_title,
                                    "source": "interactive_cta",
                                    "timestamp": timezone.now().isoformat()
                                })

                                text_body = f"[CTA Click] {button_title}"

                                print(f"🔗 CTA clicked: {button_title}")

                            else:

                                text_body = f"[Interactive] {interactive_type}"
                                mark_button_clicked(mobile)

                                button_response = json.dumps({
                                    "type": "unknown_interactive",
                                    "interactive_type": interactive_type,
                                    "raw_data": interactive,
                                    "timestamp": timezone.now().isoformat()
                                })

                                print(
                                    f"⚠️ Unknown interactive type: "
                                    f"{interactive_type}"
                                )
                        

                        elif msg_type in ("image", "video", "audio", "document"):
                            media_id = msg[msg_type].get("id")
                            content_type = msg_type
                            text_body = f"[{msg_type.title()}]"
                            # print(f"📎 {msg_type} from {mobile}, media_id: {media_id}")
                            media_file_data = download_whatsapp_media2(media_id)
                            if media_file_data:
                                pass
                                # print(f"✅ Downloaded {msg_type}")
                            else:
                                pass
                                # print(f"❌ Failed to download {msg_type}")

                        elif msg_type == "unsupported":
                            error = msg.get("errors", [{}])[0].get("message", "Unknown")
                            # print(f"⚠️ Unsupported: {error}")
                            continue
                        customer_name = ""
                        contacts_data = value.get("contacts", [])
                        if contacts_data:
                            customer_name = contacts_data[0].get("profile", {}).get("name", "")
                            print(f"📛 Customer name: {customer_name}")  # Debug print

                        last_incoming = SmsWhatsAppLog2.objects.filter(mobile=mobile,message_type='Received').order_by('-sent_at').first()
                        send_welcome = False
                        if not last_incoming:
                            send_welcome = True
                        elif (timezone.now() - last_incoming.sent_at).total_seconds() > 21600:  # 1 hour
                            send_welcome = True


                        # Save message
                        with transaction.atomic():
                            log = SmsWhatsAppLog2.objects.create(
                                customer_name=customer_name,
                                mobile=mobile,
                                template_name="incoming",
                                sent_text_message=text_body,
                                status="Unread",
                                message_type="Received",
                                message_id=msg_id,
                                content_type=content_type,
                                button_response=button_response,
                            )
                            clear_chat_cache2(mobile)


                            if media_file_data:
                                filename, content = media_file_data
                                log.media_file.save(filename, ContentFile(content))
                                log.save()
                                # print(f"💾 Saved media: {filename}")
                        
                        # ======================================
                        # AUTO LEAD CREATION FROM INTERESTED BUTTON
                        # ======================================
                        try:
                            interested_clicked = False

                            if "[Button Click] Interested" in text_body:
                                interested_clicked = True

                            if interested_clicked:
                                gm = ws_group2(mobile)
                                if gm:
                                    async_to_sync(channel_layer.group_send)(
                                        f"chat2_{gm}",
                                        {
                                            "type": "new_message",
                                            "message": {
                                                "id": log.id,
                                                "mobile": mobile,
                                                "sent_text_message": log.sent_text_message,
                                                "content_type": log.content_type,
                                                "media_file": "",
                                                "sent_at": timezone.localtime(log.sent_at).isoformat(),
                                                "message_type": "Received",
                                                "message_id": log.message_id,
                                                "status": log.status,
                                                "sender_name": customer_name
                                            }
                                        }
                                    )

                                message = (
                                    "ధన్యవాదాలు! 🙏\n\n"
                                    "మీ ఆసక్తిని నమోదు చేసుకున్నాము.\n\n"
                                    "మా Sales టీమ్ త్వరలో మిమ్మల్ని సంప్రదిస్తుంది.\n\n"
                                    "📞 8333000111\n\n"
                                    "SMSquare"
                                )

                                # ----------------------------------
                                # SEND WHATSAPP AUTO REPLY
                                # ----------------------------------
                                resp = send_whatsapp_text2(mobile, message)

                                msg_id = ""
                                try:
                                    if isinstance(resp, dict):
                                        msg_id = resp.get("messages", [{}])[0].get("id", "")
                                except Exception:
                                    pass

                                if not msg_id:
                                    msg_id = f"AUTO-{uuid.uuid4().hex[:12]}"

                                # ----------------------------------
                                # SAVE AUTO REPLY TO CHAT HISTORY
                                # ----------------------------------
                                auto_log = SmsWhatsAppLog2.objects.create(
                                    customer_name="PAdmasai",
                                    mobile=mobile,
                                    template_name="auto_reply",
                                    sent_text_message=message,
                                    status="Sent",
                                    message_type="Sent",
                                    message_id=msg_id,
                                    content_type="text",
                                )

                                clear_chat_cache2(mobile)

                                # ----------------------------------
                                # UPDATE CONTACT
                                # ----------------------------------
                                ChatContact2.objects.filter(mobile=mobile).update(
                                    last_time=timezone.now(),
                                    last_msg=message,
                                    last_type="Sent",
                                    last_status="Sent"
                                )

                                # ----------------------------------
                                # WEBSOCKET UPDATE CHAT WINDOW
                                # ----------------------------------
                                gm = ws_group2(mobile)

                                if gm:
                                    async_to_sync(channel_layer.group_send)(
                                        f"chat2_{gm}",
                                        {
                                            "type": "new_message",
                                            "message": {
                                                "id": auto_log.id,
                                                "mobile": mobile,
                                                "sent_text_message": message,
                                                "content_type": "text",
                                                "media_file": "",
                                                "sent_at": timezone.localtime(
                                                    auto_log.sent_at
                                                ).isoformat(),
                                                "message_type": "Sent",
                                                "message_id": msg_id,
                                                "status": "Sent",
                                                "sender_name": "SMSquare"
                                            }
                                        }
                                    )

                                # ----------------------------------
                                # CREATE SALES CASE
                                # ----------------------------------
                                from adminpanel.models import SupportGroup

                                existing_case = Case.objects.filter(
                                    mobile=mobile,
                                    group__name__iexact="Sales"
                                ).exclude(
                                    status__in=["Closed"]
                                ).first()

                                if not existing_case:

                                    sales_group = SupportGroup.objects.get(name="Sales")

                                    case=Case.objects.create(
                                        case_id=f"LEAD-{uuid.uuid4().hex[:8].upper()}",
                                        customer_name=customer_name,
                                        mobile=mobile,
                                        issue_description="Customer clicked Interested on WhatsApp Loan Campaign",
                                        group=sales_group,
                                        current_level="ESC2",
                                        status="Open",
                                        priority="Medium",
                                        source="WhatsApp Marketing Campaign",
                                        created_by="System Auto Lead"
                                    )
                                    case._skip_ticket_open = True
                                    case.save()

                                    print(f"✅ Sales Lead Created: {mobile}")

                        except Exception as e:
                            print("Lead creation error:", str(e))

                        # Update contact
                        obj, created = ChatContact2.objects.get_or_create(
                            mobile=mobile,
                            defaults={
                                "last_time": timezone.now(),
                                "last_msg": text_body or "",
                                "last_type": "Received",
                                "last_status": "Unread",
                                "unread": 1,
                            }
                        )
                        if not created:
                            ChatContact2.objects.filter(mobile=mobile).update(
                                last_time=timezone.now(),
                                last_msg=text_body if text_body else "[Button Click]",
                                last_type="Received",
                                last_status="Unread",
                                unread=F("unread") + 1
                            )
                        if send_welcome:
                            if was_button_clicked_recently(mobile):
                                clear_button_clicked(mobile)   # skip because button clicked
                            else:
                                send_welcome_message.delay('psf', mobile, customer_name)
                        # WebSocket broadcast
                        gm = ws_group2(mobile)
                        if gm and not interested_clicked:
                            ws_message = {
                                "id": log.id,
                                "mobile": mobile,
                                "sent_text_message": log.sent_text_message,
                                "content_type": log.content_type,
                                "media_file": log.media_file.url if log.media_file else "",
                                "sent_at": timezone.localtime(log.sent_at).isoformat(),
                                "message_type": "Received",
                                "message_id": log.message_id,
                                "status": log.status,
                                "sender_name": customer_name
                            }
                            if button_response:
                                try:
                                    btn_data = json.loads(button_response)
                                    ws_message["button_title"] = btn_data.get("button_title") or btn_data.get("list_title")
                                    ws_message["interaction_type"] = btn_data.get("type")
                                except:
                                    pass

                            async_to_sync(channel_layer.group_send)(
                                f"chat2_{gm}",
                                {"type": "new_message", "message": ws_message}
                            )

                        async_to_sync(channel_layer.group_send)(
                            "global_contacts2",
                            {
                                "type": "contact.update",
                                "contact": {
                                    "mobile": mobile,
                                    "last_msg": text_body if text_body else "[Button Click]",
                                    "last_type": "Received",
                                    "last_status": "Unread",
                                    "unread": obj.unread if created else obj.unread + 1,
                                    #"last_time": timezone.now().isoformat(),
                                }
                            }
                        )
                        if button_response:
                            print(f"✅ Button response saved for {mobile}")

                        # print(f"✅ Saved incoming {msg_type} from {mobile}")

                    # ======================================
                    # PROCESS STATUS UPDATES
                    # ======================================
                    for status in value.get("statuses", []):
                        msg_id = status.get("id")
                        status_type = (status.get("status") or "").lower()

                        if not msg_id:
                            continue

                        # print(f"📨 Status update: {msg_id} -> {status_type}")

                        # Message should already exist with real ID
                        obj = SmsWhatsAppLog2.objects.filter(message_id=msg_id).first()
                        # If not found, try to find by partial match (temp ID might still be there)
                        if not obj and len(msg_id) > 30:
                            # Try to find by the real ID pattern in any message
                            partial = msg_id[:30]
                            obj = SmsWhatsAppLog2.objects.filter(message_id__startswith=partial).first()
                            if obj:
                                # print(f"✅ Found by partial match, updating full ID")
                                SmsWhatsAppLog2.objects.filter(id=obj.id).update(message_id=msg_id)
                                obj.refresh_from_db()
                        # Also try to find by looking for messages with status "Sending"
                        if not obj:
                            # print(f"❌ Message not found: {msg_id}")
                            continue

                        mobile = obj.mobile
                        #error handling
                        errors = status.get("errors",[])

                        if status_type == "sent":
                            norm = "Sent"
                        elif status_type == "delivered":
                            norm = "Delivered"
                        elif status_type == "read":
                            norm = "Read"
                        elif status_type == "failed":
                            norm = "Failed"
                            if errors:
                                err = errors[0]
                                code = int(err.get("code", 0))
                                error_map = {
                                    131047: "24H_WINDOW_EXPIRED", 131026: "NOT_ON_WHATSAPP",
                                    131051: "UNSUPPORTED_MESSAGE_TYPE", 131011: "BLOCKED_BY_USER",
                                    130403: "BLOCKED_BY_BUSINESS", 131050: "OPTED_OUT",
                                    190: "TOKEN_ERROR", 131009: "INVALID_PARAMETER",
                                    131000: "UNKNOWN_ERROR", 131045: "REGISTRATION_ERROR",
                                    132000: "TEMPLATE_PARAM_ERROR", 132001: "TEMPLATE_NOT_FOUND",
                                    132015: "TEMPLATE_PAUSED", 132016: "TEMPLATE_DISABLED",
                                    130429: "RATE_LIMIT", 131056: "TOO_MANY_MESSAGES",
                                }
                                norm = error_map.get(code, f"Failed_{code}")

                                
                        else:
                            continue

                        # 🔥 FIXED: Update database with ALL error fields (error_code + error_reason)
                        
                        SmsWhatsAppLog2.objects.filter(message_id=msg_id).update(status=norm, error_message=json.dumps(errors) if errors else "")

                        ChatContact2.objects.filter(mobile=mobile).update(last_status=norm)

                        # WebSocket update
                        gm = ws_group2(mobile)
                        if gm:
                            async_to_sync(channel_layer.group_send)(
                                f"chat2_{gm}",
                                {
                                    "type": "delivery.update",
                                    "message_id": msg_id,
                                    "status": norm,
                                    "mobile": mobile
                                }
                            )
                        async_to_sync(channel_layer.group_send)(
                            "global_contacts2",
                            {
                                "type": "contact.update",
                                "contact": {
                                    "mobile": mobile,
                                    "last_status": norm
                                }
                            }
                        )

                        print(f"✅ Updated {msg_id} to {norm}")
                        total_unread = ChatContact2.objects.filter(unread__gt=0).count()
                        async_to_sync(channel_layer.group_send)(
                            "global_contacts2",
                            {
                                "type": "unread.update",
                                "unread_count": total_unread
                            }
                        )

                        # print(f"✅ Updated {msg_id} to {norm}")

            return JsonResponse({"status": "received"})

        except Exception as e:
            # print(f"WEBHOOK ERROR: {e}")
            traceback.print_exc()
            return JsonResponse({"error": str(e)}, status=400)

    return HttpResponseBadRequest("Unsupported method")


@csrf_exempt
def mark_read2(request, mobile):
    try:
        mobile_norm = format_mobile2(mobile)
        ChatContact2.objects.filter(mobile=mobile).update(unread=0)
        channel_layer = get_channel_layer()
        gm = ws_group2(mobile_norm)
        if gm:
            async_to_sync(channel_layer.group_send)(
                f"chat2_{gm}",
                {"type": "delivery.update", "message_id": "", "status": "Read", "mobile": mobile_norm}
            )
        async_to_sync(channel_layer.group_send)(
            "global_contacts2",
            {"type": "presence.update", "mobile": mobile_norm, "status": "updated"}
        )
        return JsonResponse({"status": "ok"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def get_contact_messages2(request):
    mobile = request.GET.get('mobile')
    if not mobile:
        return JsonResponse({'error': 'Mobile required'}, status=400)
    messages = SmsWhatsAppLog2.objects.filter(mobile=mobile).order_by('sent_at')
    data = {
        'messages': [{
            'id': m.id,
            'sent_text_message': m.sent_text_message,
            'message_type': m.message_type,
            'sent_at': m.sent_at.isoformat(),
            'content_type': m.content_type,
            'media_file': m.media_file.url if m.media_file else '',
            'status': m.status,
        } for m in messages]
    }
    return JsonResponse(data)

def view_secure_document2(request, log_id):
    if not request.session.get("messaging2_user"):
        return HttpResponseForbidden("Authentication required.")
    log = get_object_or_404(SmsWhatsAppLog2, id=log_id)
    filename = (log.media_file.name or "").lower()
    is_noc_document = "noc" in filename
    if log.content_type != "document":
        return HttpResponseForbidden("Not allowed - This is not a document")
    if log.message_type.lower() not in ['sent', 'sending']:
        return HttpResponseForbidden("Not allowed - Only sent documents can be viewed")
    if not is_noc_document:
        return HttpResponseForbidden("Not allowed - This is not a NOC document")
    if not filename.endswith('.pdf'):
        return HttpResponseForbidden("Not allowed - NOC documents must be PDF files")
    file_obj = default_storage.open(log.media_file.name, "rb")
    response = StreamingHttpResponse(file_obj, content_type="application/pdf")
    response["Content-Disposition"] = "inline; filename=NOC.pdf"
    response["Cache-Control"] = "no-store, no-cache, must-revalidate"
    return response

# =============================== Escalation views =======================================================
import logging
from django.core.cache import cache
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from financehub.models import Lcc

logger = logging.getLogger(__name__)

@csrf_exempt
def fetch_padmasai_details(request):
    mobile = request.GET.get('mobile')
    if not mobile:
        return JsonResponse({'success': False, 'error': 'Mobile number required'}, status=200)

    # Normalize: remove '+', spaces, and any non-digit characters
    clean_mobile = ''.join(filter(str.isdigit, mobile))
    # If it starts with 91, keep it; otherwise it's already without country code.

    # Generate possible variations
    possible_numbers = [clean_mobile]
    if clean_mobile.startswith('91'):
        # Also try without the leading 91
        possible_numbers.append(clean_mobile[2:])
    else:
        # Also try with 91 prefixed
        possible_numbers.append('91' + clean_mobile)

    # Optional cache (key based on original mobile)
    cache_key = f'lcc_{clean_mobile}'
    cached = cache.get(cache_key)
    if cached is not None:
        return JsonResponse({'success': True, **cached})

    try:
        lcc_record = None
        for num in possible_numbers:
            lcc_record = Lcc.objects.filter(cust_mobile=num).first()
            if lcc_record:
                break

        if not lcc_record:
            logger.info(f"❌ No LCC record for mobile {mobile} (tried: {possible_numbers})")
            return JsonResponse({'success': False, 'error': 'No details found'}, status=200)

        result = {
            'customer_name': lcc_record.customer_name or '',
            'agreement_no': lcc_record.loan_number or '',
            'vehicle_no': lcc_record.vehicle_no or '',
        }
        cache.set(cache_key, result, 300)
        return JsonResponse({'success': True, **result})

    except Exception as e:
        logger.error(f"LCC query error: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=200)


# ============================================
# ESCALATION DASHBOARDS (APP-AWARE)
# ============================================

from .models import Agent

def auto_assign(case):
    agents = Agent.objects.filter(
        groups__id=case.group.id,
        role='AGENT',
        is_active=True
    )

    agent = agents.first()

    if agent:
        case.assign_to_agent(agent, assigned_by="System")

from django.utils import timezone
from datetime import timedelta
from .models import CaseEscalationLog

from django.utils import timezone

@messaging2_required
def agent_dashboard2(request):
    CaseModel = get_case_model_for_app(request)
    agent = get_agent_from_user(request.user)

    username = agent.user.username

    today_start = timezone.localtime().replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )

    # Personal stats
    total_cases_created = CaseModel.objects.filter(
        created_by=username
    ).count()

    created_today = CaseModel.objects.filter(
        created_by=username,
        created_at__gte=today_start
    ).count()

    total_escalated = CaseEscalationLog.objects.filter(
        escalated_by=username
    ).count()

    escalated_today = CaseEscalationLog.objects.filter(
        escalated_by=username,
        created_at__gte=today_start
    ).count()

    resolved = CaseModel.objects.filter(
        resolved_by=username
    ).count()

    resolved_today = CaseModel.objects.filter(
        resolved_by=username,
        resolved_at__gte=today_start
    ).count()

    # My assigned cases only
    assigned_cases = CaseModel.objects.filter(
        assigned_to=agent,
        status__in=['Open', 'In Progress', 'Reopened']
    ).exclude(status='Closed')

    stats = {
        'total_cases_created': total_cases_created,
        'created_today': created_today,
        'total_escalated': total_escalated,
        'escalated_today': escalated_today,
        'resolved': resolved,
        'resolved_today': resolved_today,
        'my_cases': assigned_cases.count(),
        'available_cases': 0,
    }

    context = {
        'cases': assigned_cases,
        'assigned_cases': assigned_cases,
        'stats': stats,
        'agent': agent,
        'current_app': request.GET.get('app', 'psf'),
        'app_list': APP_CONFIG.items(),
    }

    return render(
        request,
        'messaging2/agent_dashboard.html',
        context
    )
@messaging2_required
def agent_case_list_api(request):
    agent = get_agent_from_user(request.user)
    CaseModel = get_case_model_for_app(request)

    username = agent.user.username
    tab = request.GET.get('tab', 'assigned')

    today_start = timezone.localtime().replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )

    if tab == 'assigned':
        cases = CaseModel.objects.filter(
            assigned_to=agent,
            status__in=['Open', 'In Progress', 'Reopened']
        ).exclude(status='Closed')

    elif tab == 'today_created':
        cases = CaseModel.objects.filter(
            created_by=username,
            created_at__gte=today_start
        ).order_by('-created_at')

    elif tab == 'resolved_by_me':
        cases = CaseModel.objects.filter(
            resolved_by=username
        ).order_by('-resolved_at')

    elif tab == 'escalated_by_me':
        cases = CaseModel.objects.filter(
            escalation_logs__escalated_by=username
        ).distinct().order_by('-created_at')

    else:
        cases = CaseModel.objects.none()

    case_list = [{
        'case_id': c.case_id,
        'customer_name': c.customer_name,
        'mobile': c.mobile,
        'loan_number': c.loan_number,
        'vehicle_number': c.vehicle_number,
        'group_name': c.group.name if c.group else None,
        'priority': c.priority,
        'status': c.status,
        'created_at': timezone.localtime(c.created_at).isoformat(),
    } for c in cases]

    return JsonResponse({'cases': case_list})

# Executive Dashboard (ESC2)


from financehub.models import Lcc

from financehub.models import Lcc
from django.db.models import Q

def normalize_mobile(mobile):
    mobile = ''.join(filter(str.isdigit, str(mobile or '')))

    if mobile.startswith('91') and len(mobile) > 10:
        mobile = mobile[-10:]

    return mobile


def enrich_case_details(cases):

    for case in cases:

        need_customer = not (
            case.customer_name and str(case.customer_name).strip()
        )

        need_loan = not (
            case.loan_number and str(case.loan_number).strip()
        )

        if not (need_customer or need_loan):
            continue

        mobile = normalize_mobile(case.mobile)

        lcc = Lcc.objects.filter(
            Q(cust_mobile__endswith=mobile) |
            Q(guarantor_mobile__endswith=mobile)
        ).first()

        if not lcc:
            continue

        updated_fields = []

        # Customer Mobile Match
        if (
            lcc.cust_mobile and
            normalize_mobile(lcc.cust_mobile) == mobile
        ):

            if need_customer and lcc.customer_name:
                case.customer_name = lcc.customer_name
                updated_fields.append('customer_name')

            if need_loan and lcc.loan_number:
                case.loan_number = lcc.loan_number
                updated_fields.append('loan_number')

        # Guarantor Mobile Match
        elif (
            lcc.guarantor_mobile and
            normalize_mobile(lcc.guarantor_mobile) == mobile
        ):

            if need_customer and lcc.guarantor:
                case.customer_name = lcc.guarantor
                updated_fields.append('customer_name')

            if need_loan and lcc.loan_number:
                case.loan_number = lcc.loan_number
                updated_fields.append('loan_number')

        if updated_fields:
            case.save(update_fields=updated_fields)

            print(
                f"UPDATED: Mobile={case.mobile}, "
                f"Name={case.customer_name}, "
                f"Loan={case.loan_number}"
            )
from adminpanel.models import SupportGroup, Subgroup,Category   # add at top if missing

@messaging2_required
def executive_dashboard2(request):
    CaseModel = get_case_model_for_app(request)
    agent = get_agent_from_user(request.user)
    if agent.role != 'EXECUTIVE':
        return redirect('agent_dashboard')

    executive_groups = agent.groups.all()
    executive_subgroups = agent.subgroup.all()

    # ─── Filters from request ──────────────────────────────────
    category_id = request.GET.get('category')
    subgroup_id = request.GET.get('subgroup')

    # ─── Build combined filter (groups/subgroups) ─────────────
    combined_filter = Q()
# Per‑group logic
    for group in executive_groups:
        group_subgroups = executive_subgroups.filter(group=group)
        if group_subgroups.exists():
            combined_filter |= Q(group=group, subgroup__in=group_subgroups)
        else:
            combined_filter |= Q(group=group)
# Also include any subgroups that may not have a group in executive_groups
    combined_filter |= Q(subgroup__in=executive_subgroups)

    if not combined_filter:
        combined_filter = Q(pk__in=[])  # no access

    # ─── Apply category filter ─────────────────────────────────
    if category_id:
        combined_filter &= Q(category_id=category_id)

    # ─── Apply subgroup filter (optional, overrides) ──────────
    if subgroup_id:
        combined_filter &= Q(subgroup_id=subgroup_id)

    # ─── Pending cases ──────────────────────────────────────────
    pending_cases = CaseModel.objects.filter(
        current_level='ESC2',
        status__in=['Open', 'In Progress', 'Reopened']
    ).exclude(status='Closed').filter(combined_filter).select_related('group', 'subgroup', 'category').order_by('-created_at')

    # ─── Resolved cases ─────────────────────────────────────────
    resolved_cases = CaseModel.objects.filter(
        resolved_at_level='ESC2',
        status='Resolved'
    ).filter(combined_filter).select_related('group', 'subgroup', 'category').order_by('-resolved_at')

    # ─── Escalated cases ────────────────────────────────────────
    escalated_cases = CaseModel.objects.filter(
        previous_level='ESC2'
    ).filter(combined_filter).select_related('group', 'subgroup', 'category').order_by('-updated_at')

    # ─── Enrich with names ──────────────────────────────────────
    enrich_case_details(pending_cases)
    enrich_case_details(resolved_cases)
    enrich_case_details(escalated_cases)


    stats = {
        'pending': pending_cases.count(),
        'resolved': resolved_cases.count(),
        'escalated': escalated_cases.count(),
    }

    current_app = request.GET.get('app', 'psf')
    agent_permissions = {
        'can_edit': agent.can_edit,
        'can_resolve': agent.can_resolve,
        'can_close': agent.can_close,
    }

    # ─── For filter dropdowns ───────────────────────────────────
    all_groups = SupportGroup.objects.all().order_by('name')
    all_subgroups = Subgroup.objects.select_related('group').order_by('group__name', 'name')
    all_categories = Category.objects.select_related('group').order_by('group__name', 'name')

    return render(request, 'messaging2/executive_dashboard.html', {
        'pending_cases': pending_cases,
        'resolved_cases': resolved_cases,
        'escalated_cases': escalated_cases,
        'stats': stats,
        'agent': agent,
        'agentPermissions': agent_permissions,
        'current_app': current_app,
        'app_list': APP_CONFIG.items(),
        'all_groups': all_groups,
        'all_subgroups': all_subgroups,
        'all_categories': all_categories,   # new
        'selected_category': category_id,
        'selected_subgroup': subgroup_id,
    })

import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.db.models import Q

def export_executive_cases(request):
    # Get agent
    try:
        agent = request.user.agent_profile
    except AttributeError:
        return HttpResponse("Agent not found", status=403)

    CaseModel = get_case_model_for_app(request)  # your dynamic function

    tab = request.GET.get('tab', 'pending')
    category_id = request.GET.get('category')
    subgroup_id = request.GET.get('subgroup')

    # Permission filter (OR)
    manager_groups = agent.groups.all()
    manager_subgroups = agent.subgroup.all()
    case_filter = Q()
    if manager_groups.exists():
        case_filter |= Q(group__in=manager_groups)
    if manager_subgroups.exists():
        case_filter |= Q(subgroup__in=manager_subgroups)
    if not case_filter:
        return HttpResponse("No permissions", status=403)

    # Base queryset based on tab
    if tab == 'pending':
        qs = CaseModel.objects.filter(
            current_level='ESC2',
            status__in=['Open', 'In Progress', 'Reopened']
        ).exclude(status='Closed')
    elif tab == 'resolved':
        qs = CaseModel.objects.filter(resolved_at_level='ESC2', status='Resolved')
    elif tab == 'escalated':
        qs = CaseModel.objects.filter(previous_level='ESC2')
    else:
        return HttpResponse("Invalid tab", status=400)

    qs = qs.filter(case_filter)
    if category_id:
        qs = qs.filter(category_id=category_id)
    if subgroup_id:
        qs = qs.filter(subgroup_id=subgroup_id)
    qs = qs.select_related('group', 'subgroup', 'category').order_by('-created_at')

    # Excel creation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab.capitalize() + " Cases"

    # Headers
    headers = ['Case ID', 'Loan No', 'Vehicle No', 'Customer', 'Mobile', 'Department', 'Subgroup', 'Category', 'Status', 'Created At']
    if tab == 'resolved':
        headers.extend(['Resolved At', 'Resolved By'])
    elif tab == 'escalated':
        headers.extend(['Current Level', 'Escalated At', 'Escalated By'])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_num, case in enumerate(qs, 2):
        row_data = [
            case.case_id,
            case.loan_number or '-',
            case.vehicle_number or '-',
            case.customer_name or case.mobile,
            case.mobile,
            case.group.name if case.group else '-',
            case.subgroup.name if case.subgroup else '-',
            case.category.name if case.category else '-',
            case.status,
            case.created_at.strftime('%Y-%m-%d %H:%M') if case.created_at else '-',
        ]
        if tab == 'resolved':
            row_data.append(case.resolved_at.strftime('%Y-%m-%d %H:%M') if case.resolved_at else '-')
            row_data.append(case.resolved_by or 'System')
        elif tab == 'escalated':
            last_log = case.escalation_logs.first()
            row_data.append(case.current_level)
            row_data.append(last_log.created_at.strftime('%Y-%m-%d %H:%M') if last_log and last_log.created_at else '-')
            row_data.append(last_log.escalated_by if last_log else 'System')

        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=val)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=executive_{tab}_cases.xlsx'
    wb.save(response)
    return response
# Manager Dashboard (ESC3)

from django.db.models import Q
from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib import messages
from messaging2.models import Agent, SupportGroup, Subgroup, Category
from .models import Case as CaseModel  # adjust import if you use a dynamic get_case_model_for_app

def manager_dashboard2(request):
    # Use the correct Agent model from messaging2 (matching login)
    try:
        agent = request.user.agent_profile
    except AttributeError:
        messages.error(request, "Agent profile not found.")
        return redirect('admin_login')

    # Role check
    if agent.role != 'MANAGER':
        return redirect('agent_dashboard')

    # Get Case model (dynamic based on app – keep your existing logic)
    CaseModel = get_case_model_for_app(request)  # assume this is defined elsewhere

    manager_groups = agent.groups.all()
    manager_subgroups = agent.subgroup.all()

    # ─── Filters from request ──────────────────────────────────
    category_id = request.GET.get('category')
    subgroup_id = request.GET.get('subgroup')

    # ─── Build combined filter (OR) for groups and subgroups ──
    case_filter = Q()
    for group in manager_groups:
        group_subgroups = manager_subgroups.filter(group=group)
        if group_subgroups.exists():
            case_filter |= Q(group=group, subgroup__in=group_subgroups)
        else:
            case_filter |= Q(group=group)
    case_filter |= Q(subgroup__in=manager_subgroups)

    # If no permissions, show empty dashboard
    if not case_filter:
        context = {
            'pending_cases': CaseModel.objects.none(),
            'resolved_cases': CaseModel.objects.none(),
            'escalated_cases': CaseModel.objects.none(),
            'closed_cases': CaseModel.objects.none(),
            'stats': {'pending': 0, 'resolved': 0, 'escalated': 0, 'closed': 0},
            'dept_stats': [],
            'level_wise_stats': [],
            'subgroup_level_stats': [],
            'agents_list': [],
            'agent': agent,
            'current_app': request.GET.get('app', 'psf'),
            'app_list': APP_CONFIG.items(),  # ensure APP_CONFIG is imported
            'all_subgroups': Subgroup.objects.select_related('group').order_by('group__name', 'name'),
            'all_categories': Category.objects.select_related('group').order_by('group__name', 'name'),
            'selected_category': None,
            'selected_subgroup': None,
        }
        return render(request, 'messaging2/manager_dashboard.html', context)

    # ─── Apply category and subgroup filters ──────────────────
    if category_id:
        case_filter &= Q(category_id=category_id)
    if subgroup_id:
        case_filter &= Q(subgroup_id=subgroup_id)

    # ─── Pending Cases (ESC3) ──────────────────────────────────
    pending_cases = (
        CaseModel.objects.filter(
            current_level='ESC3',
            status__in=['Open', 'In Progress', 'Reopened']
        )
        .exclude(status='Closed')
        .filter(case_filter)
        .select_related('group', 'subgroup', 'category')
        .order_by('-created_at')
    )

    # ─── Resolved Cases ────────────────────────────────────────
    resolved_cases = (
        CaseModel.objects.filter(
            resolved_at_level='ESC3',
            status='Resolved'
        )
        .filter(case_filter)
        .select_related('group', 'subgroup', 'category')
        .order_by('-resolved_at')
    )

    # ─── Closed Cases ──────────────────────────────────────────
    closed_cases = (
        CaseModel.objects.filter(
            status='Closed',
            resolved_at_level='ESC3'
        )
        .filter(case_filter)
        .select_related('group', 'subgroup', 'category')
        .order_by('-closed_at')
    )

    # ─── Escalated Cases ───────────────────────────────────────
    escalated_cases = (
        CaseModel.objects.filter(
            previous_level='ESC3'
        )
        .filter(case_filter)
        .select_related('group', 'subgroup', 'category')
        .order_by('-updated_at')
    )

    # ─── Enrich with names ──────────────────────────────────────
    def enrich(cases):
        for case in cases:
            case.group_name = case.group.name if case.group else None
            case.subgroup_name = case.subgroup.name if case.subgroup else None
            case.category_name = case.category.name if case.category else None

    enrich(pending_cases)
    enrich(resolved_cases)
    enrich(closed_cases)
    enrich(escalated_cases)

    stats = {
        'pending': pending_cases.count(),
        'resolved': resolved_cases.count(),
        'escalated': escalated_cases.count(),
        'closed': closed_cases.count(),
    }

    # ─── Department Wise Stats (using manager_groups directly) ─
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    dept_stats = []

    # Use the manager's groups explicitly
    groups_to_show = manager_groups  # all groups assigned to this manager

    for group in groups_to_show:
        # Pending cases for this group (respecting all filters)
        dept_pending = (
            CaseModel.objects.filter(
                group=group,
                current_level='ESC3',
                status__in=['Open', 'In Progress', 'Reopened']
            )
            .exclude(status='Closed')
            .filter(case_filter)
            .count()
        )
        # Cases created today in this group
        dept_today = (
            CaseModel.objects.filter(
                group=group,
                created_at__gte=today_start
            )
            .filter(case_filter)
            .count()
        )
        dept_stats.append({
            'name': group.name,
            'pending': dept_pending,
            'today': dept_today,
        })

    # ─── Level Wise Stats (by department) ──────────────────────
    all_levels = ['ESC1', 'ESC2', 'ESC3', 'ESC4', 'ESC5']
    level_wise_stats = []

    for group in groups_to_show:
        group_stats = {
            'group_name': group.name,
            'levels': {},
            'resolved_by_level': {},
            'total_cases': 0,
            'total_resolved': 0,
        }
        for level in all_levels:
            open_cases = (
                CaseModel.objects.filter(
                    group=group,
                    current_level=level,
                    status__in=['Open', 'In Progress', 'Reopened']
                )
                .exclude(status='Closed')
                .filter(case_filter)
                .count()
            )
            group_stats['levels'][level] = open_cases
            group_stats['total_cases'] += open_cases

            resolved = (
                CaseModel.objects.filter(
                    group=group,
                    resolved_at_level=level
                )
                .filter(case_filter)
                .count()
            )
            group_stats['resolved_by_level'][level] = resolved
            group_stats['total_resolved'] += resolved

        level_wise_stats.append(group_stats)

    # ─── Subgroup Wise Stats (NEW) ─────────────────────────────
    subgroup_level_stats = []

    # Determine which subgroups to show: either the manager's subgroups,
    # or all subgroups of the manager's groups.
    if manager_subgroups.exists():
        subgroups_to_use = manager_subgroups
    elif manager_groups.exists():
        subgroups_to_use = Subgroup.objects.filter(group__in=manager_groups)
    else:
        subgroups_to_use = Subgroup.objects.none()

    subgroups_to_use = subgroups_to_use.select_related('group').order_by('group__name', 'name')

    for subgroup in subgroups_to_use:
        base_qs = CaseModel.objects.filter(case_filter).filter(subgroup=subgroup)

        subgroup_data = {
            'subgroup_name': subgroup.name,
            'group_name': subgroup.group.name if subgroup.group else None,
            'levels': {},
            'resolved_by_level': {},
            'total_cases': 0,
            'total_resolved': 0,
        }

        for level in all_levels:
            open_count = (
                base_qs
                .filter(
                    current_level=level,
                    status__in=['Open', 'In Progress', 'Reopened']
                )
                .exclude(status='Closed')
                .count()
            )
            subgroup_data['levels'][level] = open_count
            subgroup_data['total_cases'] += open_count

            resolved_count = base_qs.filter(resolved_at_level=level).count()
            subgroup_data['resolved_by_level'][level] = resolved_count
            subgroup_data['total_resolved'] += resolved_count

        subgroup_level_stats.append(subgroup_data)

    # ─── Agents List ────────────────────────────────────────────
    if manager_subgroups.exists():
        agents_list = (
            Agent.objects.filter(
                is_active=True,
                subgroup__in=manager_subgroups
            )
            .distinct()
            .values('id', 'name', 'role')
        )
    else:
        agents_list = (
            Agent.objects.filter(
                is_active=True,
                groups__in=manager_groups
            )
            .distinct()
            .values('id', 'name', 'role')
        )

    # ─── All groups/subgroups for filters (using manager’s scope) ─
    all_groups = manager_groups  # only show groups the manager has access to

    if manager_subgroups.exists():
        all_subgroups = manager_subgroups.select_related('group').order_by('group__name', 'name')
    elif manager_groups.exists():
        all_subgroups = Subgroup.objects.filter(group__in=manager_groups).select_related('group').order_by('group__name', 'name')
    else:
        all_subgroups = Subgroup.objects.none()

    all_categories = Category.objects.select_related('group').order_by('group__name', 'name')

    context = {
        'pending_cases': pending_cases,
        'resolved_cases': resolved_cases,
        'escalated_cases': escalated_cases,
        'closed_cases': closed_cases,
        'stats': stats,
        'dept_stats': dept_stats,
        'level_wise_stats': level_wise_stats,
        'subgroup_level_stats': subgroup_level_stats,
        'agents_list': agents_list,
        'today_date': timezone.now().date().isoformat(),
        'agent': agent,
        'current_app': request.GET.get('app', 'psf'),
        'app_list': APP_CONFIG.items(),   # import APP_CONFIG from your config
        'all_groups': all_groups,
        'all_subgroups': all_subgroups,
        'all_categories': all_categories,
        'selected_category': category_id,
        'selected_subgroup': subgroup_id,
    }

    return render(request, 'messaging2/manager_dashboard.html', context)


from django.db.models import Q
from django.http import JsonResponse
from .models import Case as CaseModel  # adjust import
from messaging2.models import Agent, SupportGroup, Subgroup, Category

def manager_cases_api(request):
    # Use the correct Agent model
    try:
        agent = request.user.agent_profile
    except AttributeError:
        return JsonResponse({'error': 'Agent not found'}, status=403)

    CaseModel = get_case_model_for_app(request)  # your dynamic logic

    # ─── Filters from request ──────────────────────────────────
    department = request.GET.get('dept', 'all')
    subgroup_id = request.GET.get('subgroup', 'all')
    category_id = request.GET.get('category', 'all')

    manager_groups = agent.groups.all()
    manager_subgroups = agent.subgroup.all()

    # ─── Build permission filter (OR) ──────────────────────────
    case_filter = Q()
    if manager_groups.exists():
        case_filter |= Q(group__in=manager_groups)
    if manager_subgroups.exists():
        case_filter |= Q(subgroup__in=manager_subgroups)

    # If no permissions, return empty
    if not case_filter:
        return JsonResponse({'cases': []})

    # ─── Base queryset (ESC3 pending only) ──────────────────────
    base_qs = (
        CaseModel.objects.filter(
            current_level='ESC3',
            status__in=['Open', 'In Progress', 'Reopened']
        )
        .exclude(status='Closed')
        .filter(case_filter)
        .select_related('group', 'subgroup', 'category')
    )

    # ─── Apply department filter ────────────────────────────────
    if department != 'all':
        base_qs = base_qs.filter(group__name=department)

    # ─── Apply subgroup filter ──────────────────────────────────
    if subgroup_id != 'all':
        base_qs = base_qs.filter(subgroup_id=subgroup_id)

    # ─── Apply category filter ──────────────────────────────────
    if category_id != 'all':
        base_qs = base_qs.filter(category_id=category_id)

    # ─── Order and format ──────────────────────────────────────
    cases = base_qs.order_by('-priority', '-created_at')

    case_list = [{
        'case_id': c.case_id,
        'customer_name': c.customer_name or c.mobile,
        'mobile': c.mobile,
        'loan_number': c.loan_number or '-',
        'vehicle_number':c.vehicle_number,
        'group_name': c.group.name if c.group else None,
        'subgroup_name': c.subgroup.name if c.subgroup else None,
        'category_name': c.category.name if c.category else None,
        'priority': c.priority,
        'status': c.status,
        'created_at': c.created_at.isoformat(),
        'issue_description': c.issue_description or '',
    } for c in cases]

    return JsonResponse({'cases': case_list})
# Head Dashboard (ESC4)
from django.db.models import Q
from adminpanel.models import Subgroup
from django.db.models import Q

def build_case_permission_filter(agent):
    """
    Build a Q filter for cases based on agent's groups and subgroups.
    For each group: if agent has subgroups under it, restrict to those subgroups.
    Otherwise, include the whole group.
    Also include any subgroups that may not have a parent group in agent.groups.
    """
    filters = Q()
    head_groups = agent.groups.all()
    head_subgroups = agent.subgroup.all()

    if not head_groups and not head_subgroups:
        return Q(pk__in=[])  # no permission → empty queryset

    for group in head_groups:
        group_subgroups = head_subgroups.filter(group=group)
        if group_subgroups.exists():
            filters |= Q(group=group, subgroup__in=group_subgroups)
        else:
            filters |= Q(group=group)

    # Include subgroups even if their group is not directly assigned
    filters |= Q(subgroup__in=head_subgroups)

    return filters

@messaging2_required
def head_dashboard2(request):
    CaseModel = get_case_model_for_app(request)
    agent = get_agent_from_user(request.user)

    if agent.role != 'HEAD':
        return redirect('agent_dashboard')

    # ─── Optional filters from request ──────────────────────
    category_id = request.GET.get('category')
    subgroup_id = request.GET.get('subgroup')

    # ─── Base queryset: all ESC4 cases (no group/subgroup restriction) ──
    base_qs = CaseModel.objects.filter(
        current_level='ESC4'
    ).exclude(status='Closed').select_related('group', 'subgroup', 'category')

    # Apply optional filters if present
    if category_id:
        base_qs = base_qs.filter(category_id=category_id)
    if subgroup_id:
        base_qs = base_qs.filter(subgroup_id=subgroup_id)

    # ─── Pending (Open, In Progress, Reopened) ──────────────
    pending_qs = base_qs.filter(
        status__in=['Open', 'In Progress', 'Reopened']
    )

    # ─── Stats ──────────────────────────────────────────────────
    stats = {
        'pending': pending_qs.count(),
        'resolved': CaseModel.objects.filter(
            current_level='ESC4',   # still at ESC4 but resolved
            resolved_at_level='ESC4',
            status='Resolved'
        ).count(),
        'closed': CaseModel.objects.filter(
            current_level='ESC4',   # if we keep closed at ESC4 (or CLOSED)
            status='Closed'
        ).count(),
        'escalated_to_admin': CaseModel.objects.filter(
            current_level='ESC5'
        ).count(),
    }

    # ─── Department Stats (only for pending cases) ────────────
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    dept_stats = []
    dept_ids = pending_qs.values_list('group_id', flat=True).distinct()
    groups_with_cases = SupportGroup.objects.filter(id__in=dept_ids)

    for group in groups_with_cases:
        dept_cases = pending_qs.filter(group=group)
        dept_today = dept_cases.filter(created_at__gte=today_start).count()
        dept_stats.append({
            'name': group.name,
            'pending': dept_cases.count(),
            'today': dept_today,
        })

    # ─── Agents list (for assignment modal) ────────────────────
    agents_list = Agent.objects.filter(is_active=True).values('id', 'name', 'role')

    # ─── Enrich cases ──────────────────────────────────────────
    cases = pending_qs.order_by('-priority', '-created_at')
    for case in cases:
        case.group_name = case.group.name if case.group else None
        case.subgroup_name = case.subgroup.name if case.subgroup else None
        case.category_name = case.category.name if case.category else None

    # ─── All groups/subgroups/categories for filter dropdowns ──
    all_groups = SupportGroup.objects.all().order_by('name')
    all_subgroups = Subgroup.objects.select_related('group').order_by('group__name', 'name')
    all_categories = Category.objects.select_related('group').order_by('group__name', 'name')

    context = {
        'cases': cases,
        'stats': stats,
        'dept_stats': dept_stats,
        'agents_list': agents_list,
        'today_date': timezone.now().date().isoformat(),
        'agent': agent,
        'current_app': request.GET.get('app', 'psf'),
        'app_list': APP_CONFIG.items(),
        'all_groups': all_groups,
        'all_subgroups': all_subgroups,
        'all_categories': all_categories,
        'selected_category': category_id,
        'selected_subgroup': subgroup_id,
    }
    return render(request, 'messaging2/head_dashboard.html', context)

@messaging2_required
def head_cases_api(request):
    agent = get_agent_from_user(request.user)
    CaseModel = get_case_model_for_app(request)

    department = request.GET.get('dept', 'all')
    subgroup_id = request.GET.get('subgroup', 'all')
    category_id = request.GET.get('category', 'all')

    # Start with all ESC4 cases (pending only)
    base_qs = CaseModel.objects.filter(
        current_level='ESC4',
        status__in=['Open', 'In Progress', 'Reopened']
    ).exclude(status='Closed').select_related('group', 'subgroup', 'category')

    # Apply filters
    if department != 'all':
        base_qs = base_qs.filter(group__name=department)
    if subgroup_id != 'all':
        base_qs = base_qs.filter(subgroup_id=subgroup_id)
    if category_id != 'all':
        base_qs = base_qs.filter(category_id=category_id)

    cases = base_qs.order_by('-created_at')

    case_list = [{
        'case_id': c.case_id,
        'customer_name': c.customer_name,
        'mobile': c.mobile,
        'loan_number': c.loan_number,
        'vehicle_number': c.vehicle_number,
        'group_name': c.group.name if c.group else None,
        'subgroup_name': c.subgroup.name if c.subgroup else None,
        'category_name': c.category.name if c.category else None,
        'priority': c.priority,
        'status': c.status,
        'created_at': c.created_at.isoformat(),
    } for c in cases]

    return JsonResponse({'cases': case_list})

import json
import uuid
import re
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login
from django.contrib import messages
from django.db.models import Q, Count
from django.utils import timezone
from django.core.paginator import Paginator
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer

from .models import *

# ============================================
# AUTHENTICATION VIEWS
# ============================================

@login_required
def get_groups_api(request):
    groups = SupportGroup.objects.all().values('id', 'name')
    return JsonResponse({'groups': list(groups)})


@login_required
def get_subgroups_api(request):
    subgroups = Subgroup.objects.select_related('group').order_by('group__name', 'name')
    data = [{'id': s.id, 'name': s.name, 'group_name': s.group.name, 'group_id': s.group.id} for s in subgroups]
    return JsonResponse({'subgroups': data})




def api_categories(request):
    """
    API endpoint to fetch categories.
    Optional query param: ?group_id=<id> to filter by department.
    Returns JSON: {"categories": [{"id": 1, "name": "Bills not updated", "group_id": 3}, ...]}
    """
    group_id = request.GET.get('group_id')
    categories = Category.objects.select_related('group').all()
    if group_id:
        try:
            group_id = int(group_id)
            categories = categories.filter(group_id=group_id)
        except ValueError:
            return JsonResponse({'error': 'Invalid group_id'}, status=400)

    data = [
        {
            'id': c.id,
            'name': c.name,
            'group_id': c.group_id,
        }
        for c in categories
    ]
    return JsonResponse({'categories': data})

def get_case_detail_api2(request, case_id):
    CaseModel = get_case_model_for_app(request)
    try:
        case = CaseModel.objects.get(case_id=case_id)
        return JsonResponse({
            'success': True,
            'case': {
                'case_id': case.case_id,
                'customer_name': case.customer_name,
                'mobile': case.mobile,
                'current_level': case.current_level,
                'previous_level': case.previous_level,
                'status': case.status,
                'priority': case.priority,
                'loan_number': case.loan_number,
                'vehicle_number':case.vehicle_number,
                'assigned_to_name': case.assigned_to_name,
                'created_by': case.created_by,
                'created_at': case.created_at.isoformat(),
                'resolved_at': case.resolved_at.isoformat() if case.resolved_at else None,
                'resolved_at_level': case.resolved_at_level,
                'resolved_by_role': case.resolved_by_role,
                'resolved_by': case.resolved_by,
                'issue_description': case.issue_description,
                'resolution_notes': case.resolution_notes,
                'reopen_count': case.reopen_count,
                'group_id': case.group.id if case.group else None,
                'subgroup_id': case.subgroup.id if case.subgroup else None,
                'group_name': case.group.name if case.group else None,
                'subgroup_name': case.subgroup.name if case.subgroup else None,
                'category_id': case.category.id if case.category else None,
                'category_name': case.category.name if case.category else None,
            }
        })
    except CaseModel.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Case not found'})
    
from adminpanel.views import get_app_from_request   
def case_description_history_api(request, case_id):
    app_key = get_app_from_request(request)
    CaseModel = APP_CONFIG[app_key]['case_model']
    case = get_object_or_404(CaseModel, case_id=case_id)
    logs = case.description_logs.all().order_by('-changed_at')  # newest first
    data = [{
        'changed_at': log.changed_at.isoformat(),
        'previous_description': log.previous_description,
        'new_description': log.new_description,
        'changed_by': log.changed_by,
        'changed_by_role': log.changed_by_role,
        'level': log.level,
    } for log in logs]
    return JsonResponse({'success': True, 'logs': data})

@csrf_exempt
def escalate_case_api2(request, case_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        CaseModel, ContactModel, _, channel_group = get_models_for_app(request)
        agent = get_agent_from_user(request.user)
        case = CaseModel.objects.get(case_id=case_id)

        # Check group membership
        if not agent.can_view_case(case) and agent.role != 'ADMIN':
            return JsonResponse({'error': 'You do not have permission to view or escalate this case'}, status=403)

        data = json.loads(request.body)
        new_level = data.get('new_level')
        reason = data.get('reason', '')
        loan = data.get('loan', '')
        name = data.get('name', '')

        if not new_level:
            return JsonResponse({'error': 'New level required'}, status=400)

        # Only allow higher levels within the same department
        available_levels = case.get_available_escalation_levels()
        if new_level not in available_levels:
            return JsonResponse({'error': f'Cannot escalate from {case.current_level} to {new_level}. Allowed: {available_levels}'}, status=400)

        # Update loan and customer name if provided
        if loan:
            case.loan_number = loan
        if name:
            case.customer_name = name

        mobile = case.mobile
        case.escalate(new_level, agent, reason, loan, name)
        ContactModel.objects.filter(mobile=mobile).update(current_level=new_level)

        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            channel_group,
            {"type": "contact.update", "contact": {"mobile": case.mobile, "current_level": new_level}}
        )
        return JsonResponse({'success': True, 'message': f'Case escalated to {new_level}', 'new_level': new_level})

    except CaseModel.DoesNotExist:
        return JsonResponse({'error': 'Case not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)





from .tasks import send_ticket_close_message   # import the task

@csrf_exempt
def resolve_case_api2(request, case_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        CaseModel, ContactModel, _, channel_group = get_models_for_app(request)

        agent = get_agent_from_user(request.user)
        if not agent.has_resolve_permission():
            return JsonResponse(
                {'error': 'You do not have permission to resolve cases'},
                status=403
            )

        case = CaseModel.objects.get(case_id=case_id)

        if case.status in ['Resolved', 'Closed']:
            return JsonResponse(
                {'error': f'Case already {case.status}'},
                status=400
            )

        data = json.loads(request.body)
        resolution_notes = data.get('resolution_notes', '')

        # Use model method
        case.resolve(
            agent=agent,
            resolution_notes=resolution_notes
        )

        # Update Contact
        ContactModel.objects.filter(
            mobile=case.mobile
        ).update(
            current_level='RESOLVED'
        )

        # WebSocket Broadcast
        channel_layer = get_channel_layer()

        async_to_sync(channel_layer.group_send)(
            channel_group,
            {
                "type": "contact.update",
                "contact": {
                    "mobile": case.mobile,
                    "current_level": "RESOLVED"
                }
            }
        )

        # Send Close Message
        app_key = request.GET.get('app', 'psf')

        send_ticket_close_message.delay(
            app_key,
            case.id
        )

        return JsonResponse({
            'success': True,
            'message': 'Case resolved successfully',
            'case': {
                'case_id': case.case_id,
                'status': case.status,
                'current_level': case.current_level,
                'resolved_at_level': case.resolved_at_level,
                'resolved_by_role': case.resolved_by_role
            }
        })

    except CaseModel.DoesNotExist:
        return JsonResponse(
            {'error': 'Case not found'},
            status=404
        )

    except Exception as e:
        return JsonResponse(
            {'error': str(e)},
            status=500
        )

@csrf_exempt
def reopen_case_api2(request, case_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        CaseModel = get_case_model_for_app(request)
        agent = get_agent_from_user(request.user)
        case = CaseModel.objects.get(case_id=case_id)
        if case.status == 'Closed':
            return JsonResponse({'error': 'Cannot reopen a closed case. Only admin can reopen closed cases.'}, status=400)
        if case.status != 'Resolved':
            return JsonResponse({'error': f'Only resolved cases can be reopened. Current status: {case.status}'}, status=400)
        data = json.loads(request.body)
        reopen_reason = data.get('reopen_reason', '')
        target_level = data.get('target_level', None)
        if agent.role != 'ADMIN' and not agent.can_view_case(case):
            return JsonResponse({'error': 'You do not have permission to reopen this case'}, status=403)
        case.reopen(agent, reopen_reason, target_level)
        return JsonResponse({'success': True, 'message': f'Case reopened to {case.current_level}'})
    except CaseModel.DoesNotExist:
        return JsonResponse({'error': 'Case not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def assign_case_api2(request, case_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        CaseModel = get_case_model_for_app(request)
        agent = get_agent_from_user(request.user)
        if agent.role not in ['MANAGER', 'ADMIN']:
            return JsonResponse({'error': 'Only Manager or Admin can assign cases'}, status=403)

        case = CaseModel.objects.get(case_id=case_id)
        data = json.loads(request.body)
        target_agent_id = data.get('agent_id')
        notes = data.get('notes', '')

        target_agent = Agent.objects.get(id=target_agent_id)
        # Check if target agent belongs to the same group as the case
        if case.group not in target_agent.groups.all():
            return JsonResponse({'error': f'Agent must be a member of {case.group.name} department'}, status=400)

        case.assigned_to = target_agent
        case.assigned_to_name = target_agent.name
        case.status = 'In Progress'
        case.save()

        # Log assignment
        from .models import CaseAssignmentLog
        CaseAssignmentLog.objects.create(
            case=case,
            assigned_to=target_agent,
            assigned_by=agent.name,
            reason=notes
        )

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_case_timeline_api2(request, case_id):
    CaseModel = get_case_model_for_app(request)
    case = get_object_or_404(CaseModel, case_id=case_id)
    logs = case.escalation_logs.all()[:50]
    return JsonResponse({
        'logs': [{
            'from_level': log.from_level,
            'to_level': log.to_level,
            'escalated_by': log.escalated_by,
            'reason': log.reason,
            'created_at': log.created_at.isoformat()
        } for log in logs]
    })

def get_case_action_permissions2(request, case_id):
    try:
        CaseModel = get_case_model_for_app(request)
        agent = get_agent_from_user(request.user)
        case = CaseModel.objects.get(case_id=case_id)
        permissions = {
            'can_view': agent.can_view_case(case) or agent.role == 'ADMIN',
            'can_escalate': case.can_escalate(agent, '') and agent.role != 'ADMIN' and case.status not in ['Resolved', 'Closed'],
            'can_resolve': case.can_resolve(agent),
            'can_close': case.can_close(agent),
            'can_reopen': case.status == 'Resolved' and (agent.role == 'ADMIN' or agent.can_view_case(case)),
            'can_assign': agent.role in ['LEAD', 'ADMIN'] and case.status not in ['Resolved', 'Closed'],
        }
        escalation_options = []
        if agent.role != 'ADMIN' and case.status not in ['Resolved', 'Closed']:
            for level in agent.ESCALATION_MATRIX.get(agent.role, []):
                escalation_options.append({'level': level, 'name': dict(CaseModel.ESCALATION_CHOICES).get(level, level)})
        return JsonResponse({
            'success': True,
            'permissions': permissions,
            'escalation_options': escalation_options,
            'case_status': case.status,
            'case_level': case.current_level,
            'user_role': agent.role,
        })
    except CaseModel.DoesNotExist:
        return JsonResponse({'error': 'Case not found'}, status=404)

def get_resolved_cases_api2(request):
    CaseModel = get_case_model_for_app(request)
    agent = get_agent_from_user(request.user)
    if agent.role != 'ADMIN':
        return JsonResponse({'error': 'Admin access required'}, status=403)
    resolved_cases = CaseModel.objects.filter(status='Resolved', current_level='RESOLVED').order_by('-resolved_at')
    cases_data = [{
        'case_id': c.case_id,
        'customer_name': c.customer_name,
        'mobile': c.mobile,
        'loan_number': c.loan_number,
        'resolved_at_level': c.resolved_at_level,
        'resolved_by_role': c.resolved_by_role,
        'resolved_by': c.resolved_by,
        'resolved_at': c.resolved_at.isoformat(),
        'resolution_notes': c.resolution_notes,
        'reopen_count': c.reopen_count,
    } for c in resolved_cases]
    return JsonResponse({'success': True, 'cases': cases_data, 'total': resolved_cases.count()})

def get_dashboard_stats_api2(request):
    CaseModel = get_case_model_for_app(request)
    agent = get_agent_from_user(request.user)
    stats = {'role': agent.role, 'level': agent.level, 'name': agent.name}
    if agent.role == 'ADMIN':
        stats.update({
            'total_cases': CaseModel.objects.count(),
            'open_cases': CaseModel.objects.filter(status='Open').count(),
            'in_progress': CaseModel.objects.filter(status='In Progress').count(),
            'resolved_pending_close': CaseModel.objects.filter(status='Resolved', current_level='RESOLVED').count(),
            'closed_cases': CaseModel.objects.filter(status='Closed').count(),
            'escalated_to_admin': CaseModel.objects.filter(current_level='ESC5').count(),
        })
    elif agent.role == 'MANAGER':
        stats.update({
            'my_level_cases': CaseModel.objects.filter(current_level='ESC4', status__in=['Open', 'In Progress']).count(),
            'resolved_at_my_level': CaseModel.objects.filter(resolved_at_level='ESC4').count(),
        })
    elif agent.role == 'LEAD':
        stats.update({
            'my_level_cases': CaseModel.objects.filter(current_level='ESC3', status__in=['Open', 'In Progress']).count(),
            'resolved_at_my_level': CaseModel.objects.filter(resolved_at_level='ESC3').count(),
            'assigned_agents': Agent.objects.filter(role='AGENT', is_active=True).count(),
        })
    elif agent.role == 'LEGAL':
        stats.update({
            'my_level_cases': CaseModel.objects.filter(current_level='ESC2', status__in=['Open', 'In Progress']).count(),
            'resolved_at_my_level': CaseModel.objects.filter(resolved_at_level='ESC2').count(),
        })
    else:  # AGENT
        stats.update({
            'assigned_to_me': CaseModel.objects.filter(assigned_to=agent, status__in=['Open', 'In Progress']).count(),
            'resolved_by_me': CaseModel.objects.filter(resolved_by=agent.name).count(),
        })
    return JsonResponse(stats)

def get_case_by_mobile2(request):
    mobile = request.GET.get('mobile', '')
    if not mobile:
        return JsonResponse({'error': 'Mobile required'}, status=400)
    mobile = format_mobile2(mobile)
    CaseModel = get_case_model_for_app(request)
    case = CaseModel.objects.select_related('group', 'subgroup').filter(mobile=mobile).order_by('-created_at').first()
    if case:
        return JsonResponse({
            'case': {
                'case_id': case.case_id,
                'customer_name': case.customer_name or mobile,
                'loan_number': case.loan_number,
                'vehicle_number':case.vehicle_number,
                'group_name': case.group.name if case.group else 'No group',
                'group_id': case.group.id if case.group else None,
                'subgroup_name': case.subgroup.name if case.subgroup else None,   # NEW
                'subgroup_id': case.subgroup.id if case.subgroup else None,       # NEW
                'category_name': case.category.name if case.category else None,   # NEW
                'category_id': case.category.id if case.category else None,  
                'created_by': case.created_by or 'System',
                'assigned_to_name': case.assigned_to_name or 'Unassigned',
                'current_level': case.current_level,
                'status': case.status,
                'priority': case.priority,
                'issue_description': case.issue_description or '',
                'created_at': case.created_at.isoformat(),
            }
        })
    return JsonResponse({'case': None})


@csrf_exempt
def create_case_from_chat_api2(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        mobile = data.get('mobile', '')

        # Get app‑aware models (case, contact, log)
        CaseModel, ContactModel, LogModel, _ = get_models_for_app(request)

        # ✅ Dynamic log check – uses correct app's log table
        # if not LogModel.objects.filter(mobile=mobile).exists():
        #     app_key = request.GET.get('app', 'psf')
        #     return JsonResponse({
        #         'error': f'This number has no WhatsApp messages in the {app_key} app. Cannot create case here.'
        #     }, status=400)

        customer_name = data.get('customer_name') or mobile
        agent_name = data.get('agent_name', 'Agent')
        issue_description = data.get('issue_description', '')
        loan_number = data.get('loan_number', '')
        vehicle_number = data.get('vehicle_no', data.get('vehicle_number', ''))
        group_name = data.get('group', 'Collections')
        subgroup_id = data.get('subgroup_id', None)
        category_id = data.get('category_id', None)   # now required
        escalate_to = data.get('escalate_to', None)
        force_new = data.get('force_new', False)

        # ─── Validate group ─────────────────────────────────────────────
        group_obj = SupportGroup.objects.filter(name=group_name).first()
        if not group_obj:
            return JsonResponse({'error': 'Invalid group'}, status=400)

        # ─── Handle subgroup (optional) ────────────────────────────────
        subgroup_obj = None
        if subgroup_id:
            try:
                subgroup_obj = Subgroup.objects.get(id=subgroup_id)
                if subgroup_obj.group != group_obj:
                    return JsonResponse({'error': 'Subgroup does not belong to the selected group'}, status=400)
            except Subgroup.DoesNotExist:
                return JsonResponse({'error': 'Invalid subgroup ID'}, status=400)

        # ─── Handle category (REQUIRED) ─────────────────────────────────
        if not category_id:
            return JsonResponse({'error': 'Category is required'}, status=400)

        try:
            category_id = int(category_id)
            category_obj = Category.objects.get(id=category_id)
            if category_obj.group != group_obj:
                return JsonResponse({'error': 'Category does not belong to the selected group'}, status=400)
        except (ValueError, Category.DoesNotExist):
            return JsonResponse({'error': 'Invalid category ID'}, status=400)

        # ─── Check existing active case ────────────────────────────────
        if not force_new:
            existing_case = CaseModel.objects.filter(
                mobile=mobile,
                status__in=['Open', 'In Progress', 'Resolved']
            ).first()
            if existing_case:
                return JsonResponse({
                    'success': True,
                    'case': {
                        'case_id': existing_case.case_id,
                        'customer_name': existing_case.customer_name,
                        'assigned_to_name': existing_case.assigned_to_name or 'Unassigned',
                        'current_level': existing_case.current_level,
                        'status': existing_case.status,
                        'category': existing_case.category.name if existing_case.category else None,
                    },
                    'existing': True,
                    'message': 'An active case already exists. Create new anyway?'
                })

        # ─── Determine initial level ────────────────────────────────────
        initial_level = 'ESC1'
        if escalate_to and escalate_to.startswith('ESC') and escalate_to != 'ESC1':
            initial_level = escalate_to

        # ─── Create the case ────────────────────────────────────────────
        case_id = f"CASE-{timezone.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
        case = CaseModel.objects.create(
            case_id=case_id,
            customer_name=customer_name,
            mobile=mobile,
            loan_number=loan_number,
            vehicle_number=vehicle_number,
            issue_description=issue_description[:500],
            source='WhatsApp',
            current_level=initial_level,
            status='Open',
            priority='Medium',
            created_by=agent_name,
            group=group_obj,
            subgroup=subgroup_obj,
            category=category_obj,   # category is always set now
            assigned_to=None,
            assigned_to_name=None,
        )

        # ─── Assignment & Escalation Logic ─────────────────────────────
        if initial_level == 'ESC1':
            auto_assign(case)
            if case.assigned_to:
                case.assigned_to_name = case.assigned_to.name
                case.save(update_fields=['assigned_to_name'])
        else:
            case.escalation_logs.create(
                from_level='ESC1',
                to_level=initial_level,
                escalated_by=agent_name,
                reason='Created directly at this level'
            )
            ContactModel.objects.update_or_create(
                mobile=mobile,
                defaults={'current_level': initial_level}
            )

        ContactModel.objects.update_or_create(
            mobile=mobile,
            defaults={'current_level': case.current_level}
        )

        # ─── Log the initial description (if any) ──────────────────────
        app_key = get_app_from_request(request)
        DescriptionLogModel = APP_CONFIG[app_key]['description_log_model']
        if case.issue_description:
            
            DescriptionLogModel.objects.create(
        case=case,
        previous_description="",
        new_description=case.issue_description,
        changed_by=agent_name or "System",
        changed_by_role="System",
        level=case.current_level,  # initial level
    )
        send_ticket_open_message.delay(app_key, case.id)

        # ─── Response ────────────────────────────────────────────────────
        return JsonResponse({
            'success': True,
            'case': {
                'case_id': case.case_id,
                'customer_name': case.customer_name,
                'loan_number': case.loan_number,
                'vehicle_number': case.vehicle_number,
                'assigned_to_name': case.assigned_to_name or 'Unassigned',
                'current_level': case.current_level,
                'status': case.status,
                'subgroup_name': case.subgroup.name if case.subgroup else None,
                'category_name': case.category.name if case.category else None,
            },
            'existing': False
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())   # for debugging
        return JsonResponse({'error': str(e)}, status=500)

    
@login_required
def get_description_history_api(request, case_id):
    app_key = get_app_from_request(request)
    CaseModel = APP_CONFIG[app_key]['case_model']
    case = get_object_or_404(CaseModel, case_id=case_id)

    # Permission check (optional, but recommended)
    agent = get_agent_from_user(request.user)
   

    logs = case.description_logs.all().order_by('changed_at')
    data = [{
        'description': log.new_description,          # ✅ Use new_description
        'updated_by': log.changed_by,                # ✅ Use changed_by
        'level': log.level,
        'created_at': log.changed_at.isoformat(),    # ✅ Use changed_at
    } for log in logs]

    return JsonResponse({'success': True, 'logs': data})

@messaging2_required
def get_user_role_api2(request):
    try:
        agent = Agent.objects.get(user=request.user)
        role_display_names = {'AGENT': 'Normal Agent', 'LEGAL': 'Legal Team', 'LEAD': 'Team Lead', 'MANAGER': 'Manager', 'ADMIN': 'Administrator'}
        esc_levels = {'ESC1': {'icon': '1', 'color': '#4caf50', 'name': 'Level 1 - Agent'}, 'ESC2': {'icon': '2', 'color': '#2196f3', 'name': 'Level 2 - Legal'}, 'ESC3': {'icon': '3', 'color': '#ff9800', 'name': 'Level 3 - Lead'}, 'ESC4': {'icon': '4', 'color': '#9c27b0', 'name': 'Level 4 - Manager'}, 'ESC5': {'icon': '5', 'color': '#f44336', 'name': 'Level 5 - Admin'}}
        level = agent.level
        level_info = esc_levels.get(level, {'icon': '?', 'color': '#666', 'name': 'Unknown'})
        return JsonResponse({'success': True, 'role': agent.role, 'role_display': role_display_names.get(agent.role, agent.role), 'level': level, 'level_info': level_info, 'name': agent.name, 'email': agent.email})
    except Agent.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Agent profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone

@messaging2_required
def export_cases_excel(request):
    """Export cases from the current agent tab (assigned, today_created, resolved_by_me, escalated_by_me)"""
    agent = get_agent_from_user(request.user)
    CaseModel = get_case_model_for_app(request)
    tab = request.GET.get('tab', 'assigned')

    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    base_qs = CaseModel.objects.filter(group__in=agent.groups.all())

    if tab == 'assigned':
        cases = base_qs.filter(assigned_to=agent, current_level='ESC1', status__in=['Open','In Progress','Reopened']).exclude(status='Closed')
    elif tab == 'today_created':
        cases = base_qs.filter(created_at__gte=today_start)
    elif tab == 'resolved_by_me':
        cases = CaseModel.objects.filter(resolved_by=agent.name).order_by('-resolved_at')
    elif tab == 'escalated_by_me':
        escalated_case_ids = CaseEscalationLog.objects.filter(escalated_by=agent.name).values_list('case_id', flat=True).distinct()
        cases = CaseModel.objects.filter(id__in=escalated_case_ids).order_by('-created_at')
    else:
        cases = base_qs.none()

    data = []
    for case in cases:
        data.append({
            'Case ID': case.case_id,
            'Customer Name': case.customer_name or '',
            'Mobile': case.mobile,
            'Loan Number': case.loan_number or '',
            'Department': case.group.name if case.group else '',
            'Status': case.status,
            'Created At': case.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Cases')
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="cases_{tab}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response


@messaging2_required
def export_group_cases_excel(request):
    """Export all cases from a specific group (department) - NO RESTRICTIONS"""
    agent = get_agent_from_user(request.user)
    group_id = request.GET.get('group_id')
    CaseModel = get_case_model_for_app(request)

    if not group_id:
        return JsonResponse({'error': 'Group ID required'}, status=400)

    # ✅ No group membership check – anyone can export any group
    if group_id == 'all':
        cases = CaseModel.objects.filter(group__isnull=False).order_by('-created_at')
        group_name = 'All_Departments'
    else:
        # Get the group name even if agent is not a member
        try:
            group = SupportGroup.objects.get(id=group_id)
            group_name = group.name
        except SupportGroup.DoesNotExist:
            return JsonResponse({'error': 'Group not found'}, status=404)
        cases = CaseModel.objects.filter(group_id=group_id).order_by('-created_at')

    data = []
    for case in cases:
        data.append({
            'Case ID': case.case_id,
            'Customer Name': case.customer_name or '',
            'Mobile': case.mobile,
            'Loan Number': case.loan_number or '',
            'Department': case.group.name if case.group else '',
            'Current Level': case.current_level,
            'Status': case.status,
            'Created At': case.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })

    import pandas as pd
    from io import BytesIO
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Group Cases')
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{group_name}_all_cases_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response

@csrf_exempt
def close_case_api2(request, case_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        CaseModel, ContactModel, _, channel_group = get_models_for_app(request)
        agent = get_agent_from_user(request.user)
        if not agent.has_close_permission():
            return JsonResponse(
                {'error': 'You do not have permission to close cases'},
                status=403
            )

        case = CaseModel.objects.get(case_id=case_id)
        if not case.can_close(agent):
            return JsonResponse({'error': f'Cannot close case in {case.status} status'}, status=400)
        data = json.loads(request.body)
        close_reason = data.get('close_reason', '')
        case.close(agent, close_reason)
        ContactModel.objects.filter(mobile=case.mobile).update(current_level='CLOSED', last_status='Closed')
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            channel_group,
            {"type": "contact.update", "contact": {"mobile": case.mobile, "current_level": 'CLOSED'}}
        )
        return JsonResponse({'success': True, 'message': f'Case {case.case_id} closed successfully'})
    except CaseModel.DoesNotExist:
        return JsonResponse({'error': 'Case not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
