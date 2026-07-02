from celery import shared_task
import openpyxl
import pandas as pd
import os

from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.apps import apps

from .models import SPLUploadHistory

BULK_BATCH_SIZE = 2000
PANDAS_CHUNK_SIZE = 5000

SMART_HEADER_MAP = {

    # ================= BASIC =================
    "company": "company",
    "branch": "branch",
    "centre": "centre",

    "loan_no": "loan_no",
    "loan no": "loan_no",

    "vehicleno": "vehicle_no",
    "vehicle no": "vehicle_no",

    "cif id": "cif_id",

    # ================= CUSTOMER =================
    "customer name": "customer_name",
    "customer father name": "customer_father_name",
    "customer address": "customer_address",

    # ================= GUARANTOR =================
    "guarantor name": "guarantor_name",
    "guarantor father name": "guarantor_father_name",
    "guarantor mobile num": "guarantor_mobile",
    "guarantor address": "guarantor_address",

    # ================= CO BORROWER =================
    "co borrower name": "co_borrower_name",
    "co borrower father name": "co_borrower_father_name",
    "co borrower mobile number": "co_borrower_mobile",
    "co borrower address": "co_borrower_address",

    # ================= VEHICLE =================
    "make": "make",
    "class": "vehicle_class",
    "variant": "variant",
    "vehicle type": "vehicle_type",

    "engine no": "engine_no",
    "chassis no": "chassis_no",

    "fuel type": "fuel_type",

    # ================= LOAN =================
    "loan date": "loan_date",
    "loan amount": "loan_amount",
    "tenure": "tenure",

    "loan closure date": "loan_closure_date",
    "maturity date": "maturity_date",

    "type": "loan_type",
    "reason": "reason",
    "remarks": "remarks",

    # ================= FINANCIAL =================
    "waiver": "waiver",
    "finance amount": "finance_amount",
    "installment received amount": "installment_received_amount",
    "loan closure amount": "loan_closure_amount",
    "difference amount": "difference_amount",
    "total": "total",
    "irr": "irr",
    "amount": "amount",

    # ================= NOC =================
    "noc issued to": "noc_issued_to",
    "noc date": "noc_date",

    # ================= CLASSIFICATION =================
    "loan segment": "loan_segment",
    "scheme name": "scheme_name",
    "source name": "source_name",

    # ================= COLLECTION =================
    "received installments": "received_installments",

    "principle portion collected": "principal_collected",  # 🔥 spelling fix
    "interest portion collected": "interest_collected",
    "broken interest collected": "broken_interest_collected",

    "vas charges collected": "vas_charges_collected",

    "vas collect later received": "vas_collect_later_received",

    # ================= APPROVAL =================
    "cust number": "customer_mobile",
    "cust_number": "customer_mobile",

    # 🔥 principal outstanding
    "principal_outstanding": "principal_outstanding",

    # 🔥 final approval date
    "final_approval_date": "final_approval_date",

    # ================= OUTSTANDING =================
    "interest outstanding": "interest_outstanding",
    "broken interest outstanding": "broken_interest_outstanding",

    "foreclosure charges": "foreclosure_charges",
    "foreclosure charges tax": "foreclosure_charges_tax",

    "vas charges outstanding": "vas_charges_outstanding",
    "lpc outstanding": "lpc_outstanding",
    "vas collect later oustanding": "vas_collect_later_outstanding",  # 🔥 typo fix

    # ================= WAIVER =================
    "principal bad debt": "principal_bad_debt",

    "interest waiver": "interest_waiver",
    "broken interest waiver": "broken_interest_waiver",
    "vas charges waiver": "vas_charges_waiver",
    "lpc waiver": "lpc_waiver",
    "vas collect later waiver": "vas_collect_later_waiver",

}

# ---------------------------------------------------------
# HEADER CLEANER (FIXED)
# ---------------------------------------------------------
# def clean_header(header):
#     if not header:
#         return ""

#     raw = str(header).strip().lower()

#     # normalize multiple spaces
#     raw = " ".join(raw.split())

#     # version 1: exact raw
#     if raw in SMART_HEADER_MAP:
#         return SMART_HEADER_MAP[raw]

#     # version 2: remove dots
#     no_dot = raw.replace(".", "")
#     if no_dot in SMART_HEADER_MAP:
#         return SMART_HEADER_MAP[no_dot]

#     # version 3: underscore format
#     cleaned = (
#         raw
#         .replace(".", "")
#         .replace(" ", "_")
#         .replace("-", "_")
#     )
#     if cleaned in SMART_HEADER_MAP:
#         return SMART_HEADER_MAP[cleaned]

#     # version 4: compact (no space, no underscore)
#     compact = cleaned.replace("_", "")
#     if compact in SMART_HEADER_MAP:
#         return SMART_HEADER_MAP[compact]

#     return cleaned

def clean_header(header):
    if not header:
        return ""

    raw = str(header).strip().lower()

    # normalize spaces
    raw = " ".join(raw.split())

    # 🔥 generate all possible variations
    variations = [
        raw,                                   # original
        raw.replace(".", ""),                  # remove dots
        raw.replace(".", "").replace(" ", "_"),# underscore
        raw.replace(" ", "_"),                 # only underscore
        raw.replace("-", "_"),                 # dash fix
    ]

    # compact version (no spaces / underscores)
    variations.append(
        raw.replace(".", "").replace(" ", "").replace("_", "")
    )

    # 🔥 try all variations
    for v in variations:
        if v in SMART_HEADER_MAP:
            return SMART_HEADER_MAP[v]

    return variations[2]  # default cleaned

# ---------------------------------------------------------
# SAFE DECIMAL PARSER
# ---------------------------------------------------------
from decimal import Decimal, InvalidOperation

def parse_decimal_safe(value):
    if value is None:
        return None

    value = str(value).strip().lower()

    # 🔥 handle all N/A variations
    if value in ("", "nan", "n/a", "#n/a", "na", ".", "-"):
        return None

    try:
        # remove currency
        value = value.replace("₹", "").replace("$", "")

        # remove commas
        value = value.replace(",", "")

        # remove brackets
        if value.startswith("(") and value.endswith(")"):
            value = value[1:-1].strip()

        return Decimal(value)

    except (InvalidOperation, ValueError):
        return None

# ---------------------------------------------------------
# SAFE DATE PARSER
# ---------------------------------------------------------
import pandas as pd

def parse_date_safe(value, dayfirst=True):
    # Step 0: handle actual NaN values
    if pd.isna(value):
        return None

    value_str = str(value).strip().lower()
    if value_str in ("", "nan", "n/a", "#n/a", "na", "-"):
        return None

    try:
        dt = pd.to_datetime(value, dayfirst=dayfirst, errors="coerce")
        if pd.isna(dt):
            return None
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return None


# ---------------------------------------------------------
# STRING CLEANER
# ---------------------------------------------------------
def clean_string(val):
    if val in ("", "nan", "n/a", "#n/a", "na",None,"-"):
        return "-"
    return str(val).strip()


# ---------------------------------------------------------
# MODEL RESOLVER
# ---------------------------------------------------------
def get_model_by_type(file_type: str):
    mapping = {
        "write_off": "Write_Off",
        "dealer_ta_balances": "Dealer_TA_Balances",
        "auction": "Auction",
        "ledger": "Ledger"
    }

    model_name = mapping.get(file_type.lower())
    if not model_name:
        return None

    return apps.get_model("special_cases", model_name)


# ---------------------------------------------------------
# UNIQUE FIELD DETECTOR
# ---------------------------------------------------------
def get_unique_field(model_fields):
    if "loan_no" in model_fields:
        return "loan_no"
    if "employee_id" in model_fields:
        return "employee_id"
    # if "sl_no" in model_fields:
    #     return "sl_no"
    return None


# ---------------------------------------------------------
# UNIVERSAL FILE PROCESSOR
# ---------------------------------------------------------
@shared_task(bind=True)
def process_universal_file(self, upload_id, tmp_path, ext, file_type):

    upload = SPLUploadHistory.objects.get(id=upload_id)

    try:
        if upload.status == "completed":
            return

        upload.status = "processing"
        upload.save(update_fields=["status"])

        Model = get_model_by_type(file_type)
        if not Model:
            upload.status = "error"
            upload.error_message = f"Invalid file_type: {file_type}"
            upload.save()
            return

        model_fields = {f.name for f in Model._meta.fields}
        unique_field = get_unique_field(model_fields)

        existing_values = set()
        if unique_field:
            existing_values = set(
                Model.objects.values_list(unique_field, flat=True)[:500000]
            )

        processed_rows = 0

        # =====================================================
        # ====================== CSV ===========================
        # =====================================================
        if ext == "csv":

            def process_chunk(chunk):
                nonlocal processed_rows, existing_values

                headers = [clean_header(h) for h in chunk.columns]
                chunk.columns = headers

                upload.total_rows += len(chunk)
                upload.save(update_fields=["total_rows"])

                batch = []

                for row in chunk.to_dict("records"):

                    cleaned = {}

                    for col, val in row.items():
                        if col not in model_fields:
                            continue

                        field = Model._meta.get_field(col)
                        field_type = field.get_internal_type()

                        if isinstance(val, str):
                            val = val.strip()

                        if field_type == "DateField":
                            cleaned[col] = parse_date_safe(val)

                        elif field_type in ["DecimalField", "FloatField"]:
                            cleaned[col] = parse_decimal_safe(val)

                        elif field_type == "IntegerField":
                            try:
                                cleaned[col] = int(str(val).replace(",", ""))
                            except:
                                cleaned[col] = None

                        else:
                            cleaned[col] = clean_string(val)

                    if not cleaned:
                        continue

                    if unique_field:
                        val = cleaned.get(unique_field)
                        if not val or val in existing_values:
                            continue
                        existing_values.add(val)

                    batch.append(Model(**cleaned))
                    processed_rows += 1

                    if len(batch) >= BULK_BATCH_SIZE:
                        with transaction.atomic():
                            Model.objects.bulk_create(batch, ignore_conflicts=True)
                        batch = []

                if batch:
                    with transaction.atomic():
                        Model.objects.bulk_create(batch, ignore_conflicts=True)

                upload.processed_rows = processed_rows
                upload.save(update_fields=["processed_rows"])

            try:
                reader = pd.read_csv(
                    tmp_path,
                    dtype=str,
                    keep_default_na=False,
                    chunksize=PANDAS_CHUNK_SIZE,
                    encoding="utf-8"
                )
            except:
                reader = pd.read_csv(
                    tmp_path,
                    dtype=str,
                    keep_default_na=False,
                    chunksize=PANDAS_CHUNK_SIZE,
                    encoding="latin-1"
                )

            for chunk in reader:
                process_chunk(chunk)

        # =====================================================
        # ==================== EXCEL ===========================
        # =====================================================
        elif ext in ("xlsx", "xls"):

            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb.active

            raw_headers = next(ws.iter_rows(values_only=True))
            headers = [clean_header(h) for h in raw_headers]

            upload.total_rows = max(ws.max_row - 1, 0)
            upload.save(update_fields=["total_rows"])

            batch = []

            for row in ws.iter_rows(min_row=2, values_only=True):

                if not row or all(v in (None, "", " ") for v in row):
                    continue

                row_dict = dict(zip(headers, row))
                cleaned = {}

                for col, val in row_dict.items():
                    if col not in model_fields:
                        continue

                    field = Model._meta.get_field(col)
                    field_type = field.get_internal_type()

                    if isinstance(val, str):
                        val = val.strip()

                    if field_type == "DateField":
                        cleaned[col] = parse_date_safe(val)

                    elif field_type in ["DecimalField", "FloatField"]:
                        cleaned[col] = parse_decimal_safe(val)

                    elif field_type == "IntegerField":

                        try:
                            cleaned[col] = int(str(val).replace(",", ""))
                        except:
                            cleaned[col] = None

                    else:
                        cleaned[col] = clean_string(val)

                if not cleaned:
                    continue

                if unique_field:
                    val = cleaned.get(unique_field)
                    if not val or val in existing_values:
                        continue
                    existing_values.add(val)

                batch.append(Model(**cleaned))
                processed_rows += 1

                if len(batch) >= BULK_BATCH_SIZE:
                    with transaction.atomic():
                        Model.objects.bulk_create(batch, ignore_conflicts=True)
                    batch = []

            if batch:
                with transaction.atomic():
                    Model.objects.bulk_create(batch, ignore_conflicts=True)

            upload.processed_rows = processed_rows
            upload.save(update_fields=["processed_rows"])

        upload.status = "completed"
        upload.save(update_fields=["status"])

    except Exception as e:
        upload.status = "error"
        upload.error_message = str(e)
        upload.save()

    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass



# ================================Chat App ========================================
import io
import time
import logging
from math import ceil
from django.core.files.base import ContentFile
from .utils import open_legal_pdf3

import pandas as pd
import requests
from celery import shared_task
from celery.utils.log import get_task_logger
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from django.conf import settings
from django.utils import timezone
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import transaction, close_old_connections
from django.db.models import F

from .models import *
from .utils import *

logger = get_task_logger(__name__)
logger.setLevel(logging.INFO)

# ==================================================
# MAIN BULK JOB
# ==================================================

def make_session():
    s = requests.Session()
    retries = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=(429, 500, 502, 503, 504),
    )
    s.mount("https://", HTTPAdapter(max_retries=retries))
    s.headers.update({"Content-Type": "application/json"})
    return s

def upload_legal_pdf_to_whatsapp3(pdf_filename, folder):
    """Upload PDF to WhatsApp - handles bytes correctly"""
    from io import BytesIO

    # Get PDF bytes from S3 or local
    pdf_bytes = open_legal_pdf3(pdf_filename, folder)

    if not pdf_bytes:
        raise ValueError(f"Empty PDF: {pdf_filename}")

    # Create a BytesIO object that mimics a file
    file_obj = BytesIO(pdf_bytes)
    file_obj.name = pdf_filename
    file_obj.content_type = "application/pdf"

    return upload_whatsapp_media3(file_obj)

@shared_task(bind=True, queue="special_cases")
def process_bulk_whatsapp3(self, excel_s3_path, template_choice, job_id, chunk_size=50):
    template_choice = str(template_choice)
    close_old_connections()

    try:
        job = BulkJob3.objects.get(job_id=job_id)
    except BulkJob3.DoesNotExist:
        return

    if job.status != "Pending":
        return

    job.status = "Running"
    job.started_at = timezone.now()
    job.save(update_fields=["status", "started_at"])

    # Read Excel
    try:
        with default_storage.open(excel_s3_path, "rb") as f:
            data = f.read()
        df = pd.read_excel(io.BytesIO(data), dtype=str).fillna("")
    except Exception:
        job.status = "Failed"
        job.completed_at = timezone.now()
        job.save(update_fields=["status", "completed_at"])
        return

    rows = df.to_dict("records")
    total = len(rows)

    # Total customers
    # if template_choice == "17":
    #     job.total_customers = len({
    #         format_mobile3(r.get("cust_mobile") or r.get("CustMobile"))
    #         for r in rows
    #         if r.get("cust_mobile") or r.get("CustMobile")
    #     })
    # else:
    job.total_customers = total

    job.save(update_fields=["total_customers"])

    if total == 0:
        job.status = "Completed"
        job.completed_at = timezone.now()
        job.save(update_fields=["status", "completed_at"])
        return

    # Create batches
    for i in range(0, total, chunk_size):
        process_bulk_whatsapp_batch3.apply_async(
            args=(excel_s3_path, template_choice, job_id, i, min(i + chunk_size, total)),
            queue="special_cases",
        )

@shared_task(bind=True, queue="special_cases")
def process_bulk_whatsapp_batch3(self, excel_s3_path, template_choice, job_id, start, end):
    from django.db import close_old_connections
    from django.core.files.storage import default_storage
    from django.db.models import F
    from django.utils import timezone
    import io
    import pandas as pd
    import re
    import requests

    close_old_connections()
    template_choice = str(template_choice)

    try:
        job = BulkJob3.objects.get(job_id=job_id)
    except BulkJob3.DoesNotExist:
        return

    if job.status != "Running":
        return

    # Read Excel chunk
    with default_storage.open(excel_s3_path, "rb") as f:
        df = pd.read_excel(io.BytesIO(f.read()), dtype=str).fillna("")

    rows = df.to_dict("records")[start:end]
    local_success = 0
    local_failed = 0

    # Setup session
    session = requests.Session()
    session.headers.update({
        "Authorization": f"Bearer {settings.WHATSAPP3_ACCESS_TOKEN}",
        "Content-Type": "application/json",
    })
    post_url = f"https://graph.facebook.com/v22.0/{settings.WHATSAPP3_PHONE_NUMBER_ID}/messages"

    for row in rows:
        name = row.get("customer_name") or row.get("CustomerName") or ""
        mobile = format_mobile3(row.get("cust_mobile") or row.get("CustMobile") or "")

        print(f"📱 Sending to {mobile} without number check")

        try:
            # Build payload – always sends "wel"
            payload, rendered_text = build_payload3(template_choice, row)

            # Send the request
            resp = session.post(post_url, json=payload, timeout=30)
            if not resp.ok:
                raise ValueError(f"API Error: {resp.text}")

            msg_id = resp.json()["messages"][0]["id"]

            # Log success
            log = SmsWhatsAppLog3.objects.create(
                job_id=job_id,
                customer_name=name,
                mobile=mobile,
                template_name=template_choice,   # store the choice, not the actual template name
                sent_text_message=rendered_text,
                status="Sent",
                message_id=msg_id,
                message_type="Sent",
                content_type="text",   # no media
            )

            # Update contact
            ChatContact3.objects.update_or_create(
                mobile=mobile,
                defaults={
                    "last_msg": rendered_text,
                    "last_time": timezone.now(),
                    "last_type": "Sent",
                    "last_status": "Sent",
                    "unread": 0
                }
            )

            # WebSocket broadcast (if you need it)
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            channel_layer = get_channel_layer()
            gm = re.sub(r"\D", "", mobile)
            if gm:
                async_to_sync(channel_layer.group_send)(
                    f"chat3_{gm}",
                    {
                        "type": "new_message",
                        "message": {
                            "id": log.id,
                            "mobile": mobile,
                            "sent_text_message": rendered_text,
                            "content_type": "text",
                            "media_file": "",
                            "sent_at": log.sent_at.isoformat(),
                            "message_type": "Sent",
                            "message_id": msg_id,
                            "status": "Sent",
                            "sender_name": name,
                        }
                    }
                )
                async_to_sync(channel_layer.group_send)(
                    "global_contacts3",
                    {
                        "type": "contact.update",
                        "contact": {
                            "mobile": mobile,
                            "last_msg": rendered_text,
                            "last_time": timezone.now().isoformat(),
                            "last_type": "Sent",
                            "last_status": "Sent",
                            "unread": 0
                        }
                    }
                )

            local_success += 1
            print(f"✅ Successfully sent to {mobile} - Message ID: {msg_id}")

        except Exception as e:
            err = str(e)
            print(f"❌ Failed for {mobile}: {err}")

            # Get name again (if needed)
            name = ""
            for r in rows:
                if (format_mobile3(r.get("cust_mobile") or r.get("CustMobile") or "")) == mobile:
                    name = r.get("customer_name") or r.get("CustomerName") or ""
                    break

            # Map known error codes (optional)
            status_value = "Failed"
            ERROR_MAP = {
                "131026": "NOT_ON_WHATSAPP",
                "131011": "BLOCKED_BY_USER",
                "130403": "BLOCKED_BY_BUSINESS",
                "131050": "OPTED_OUT",
                "190": "TOKEN_ERROR",
                "131009": "INVALID_PARAMETER",
                "131000": "UNKNOWN_ERROR",
                "131045": "REGISTRATION_ERROR",
                "131047": "24H_WINDOW_EXPIRED",
                "131051": "UNSUPPORTED_MESSAGE_TYPE",
                "132000": "TEMPLATE_PARAM_ERROR",
                "132001": "TEMPLATE_NOT_FOUND",
                "132015": "TEMPLATE_PAUSED",
                "132016": "TEMPLATE_DISABLED",
                "130429": "RATE_LIMIT",
                "131056": "TOO_MANY_MESSAGES",
            }
            for code, label in ERROR_MAP.items():
                if code in err:
                    status_value = label
                    break

            SmsWhatsAppLog3.objects.create(
                job_id=job_id,
                customer_name=name,
                mobile=mobile,
                template_name=template_choice,
                status=status_value,
                message_type="Failed",
                error_message=err,
            )
            local_failed += 1

    # Update job progress
    BulkJob3.objects.filter(job_id=job_id).update(
        sent_count=F("sent_count") + (local_success + local_failed),
        success_count=F("success_count") + local_success,
        failed_count=F("failed_count") + local_failed,
    )

    job.refresh_from_db()
    if job.sent_count >= job.total_customers:
        job.status = "Completed"
        job.completed_at = timezone.now()
        job.save(update_fields=["status", "completed_at"])
        finalize_bulk_job3.delay(job_id)
# FINALIZER
# ==================================================
@shared_task(bind=True, queue="special_cases")
def finalize_bulk_job3(self, job_id):
    try:
        job = BulkJob3.objects.get(job_id=job_id)
    except BulkJob3.DoesNotExist:
        return

    # Prevent double execution
    if job.success_report and job.failed_report:
        return

    # Get all field names from SmsWhatsAppLog (avoid auto-created ones like id, but include them anyway)
    from django.apps import apps
    model = SmsWhatsAppLog3
    field_names = [f.name for f in model._meta.get_fields() if not f.auto_created]
    # For better readability, you can also define a fixed list of columns you want in reports
    # field_names = ['id', 'job_id', 'customer_name', 'mobile', 'template_name', 'status', 'sent_text_message', 'error_message', ...]

    success_qs = SmsWhatsAppLog3.objects.filter(
        job_id=job_id, status__in=["Sent", "Delivered"]
    )
    failed_qs = SmsWhatsAppLog3.objects.filter(
        job_id=job_id, status="Failed"
    )

    # Build DataFrames – always with correct columns
    if success_qs.exists():
        success_df = pd.DataFrame(list(success_qs.values()))
    else:
        success_df = pd.DataFrame(columns=field_names)

    if failed_qs.exists():
        failed_df = pd.DataFrame(list(failed_qs.values()))
    else:
        failed_df = pd.DataFrame(columns=field_names)

    # 🔥 Remove timezone from datetime columns (same as before)
    for df in [success_df, failed_df]:
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.tz_localize(None)

    success_path = f"reports3/{job_id}_success.xlsx"
    failed_path = f"reports3/{job_id}_failed.xlsx"

    success_buffer = io.BytesIO()
    failed_buffer = io.BytesIO()

    # Always write both files (even empty ones)
    success_df.to_excel(success_buffer, index=False)
    failed_df.to_excel(failed_buffer, index=False)

    # Delete old files if they exist (optional but safe)
    if default_storage.exists(success_path):
        default_storage.delete(success_path)
    if default_storage.exists(failed_path):
        default_storage.delete(failed_path)

    # Save both reports unconditionally
    default_storage.save(success_path, ContentFile(success_buffer.getvalue()))
    job.success_report = success_path

    default_storage.save(failed_path, ContentFile(failed_buffer.getvalue()))
    job.failed_report = failed_path

    job.status = "Completed"
    job.completed_at = timezone.now()
    job.save(update_fields=[
        "success_report", "failed_report", "status", "completed_at"
    ])

    logger.info("Job %s COMPLETED and both reports generated", job_id)

@shared_task(queue="special_cases")
def process_pending_webhook_updates3():
    """Process any status updates that arrived before the message was saved"""
    from django.core.cache import cache
    from django.utils import timezone
    from datetime import timedelta

    keys = cache.keys("pending_wa_status_*")

    for key in keys:
        data = cache.get(key)
        if not data:
            continue

        msg_id = key.replace("pending_wa_status_", "")
        status_type = data.get('status')

        # Try to find the message now
        obj = SmsWhatsAppLog3.objects.filter(message_id=msg_id).first()

        if obj:
            if status_type == "sent":
                norm = "Sent"
            elif status_type == "delivered":
                norm = "Delivered"
            elif status_type == "read":
                norm = "Read"
            else:
                continue

            SmsWhatsAppLog3.objects.filter(id=obj.id).update(status=norm)
            print(f"✅ Processed pending status for {msg_id} -> {norm}")
            cache.delete(key)
        else:
            # If older than 60 seconds, remove
            timestamp = data.get('timestamp')
            if timestamp:
                from dateutil import parser
                if parser.parse(timestamp) < timezone.now() - timedelta(seconds=60):
                    cache.delete(key)

